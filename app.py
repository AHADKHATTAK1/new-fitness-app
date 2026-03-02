from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from flask_compress import Compress
from werkzeug.utils import secure_filename
from models import Gym, Member, Fee, MemberNote, User, StaffAccess, get_session
from gym_manager import GymManager
from auth_manager import AuthManager
from payment_manager import PaymentManager
from email_utils import EmailSender
import os
import json
from datetime import datetime, timedelta
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from io import BytesIO
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
import qrcode
import base64
from dotenv import load_dotenv
from google_wallet import GymWalletPass
import stripe
import secrets
import traceback
import threading
from subscription_tiers import TIERS, TIERS_PAKISTAN, TierManager
from tier_routes import init_upgrade_routes
from automation_manager import AutomationManager

try:
    from apscheduler.schedulers.background import BackgroundScheduler
except Exception as scheduler_import_error:
    BackgroundScheduler = None
    if os.getenv('APP_VERBOSE_STARTUP', '0') == '1':
        print(f"⚠️ APScheduler unavailable, background jobs disabled: {scheduler_import_error}")
    else:
        print("⚠️ APScheduler unavailable, background jobs disabled")

# Load environment variables from .env file
load_dotenv()

print("=" * 80)
print("🚀 STARTING GYM MANAGER APPLICATION")
print("=" * 80)

# Initialize database tables on startup
try:
    from models import init_db
    print("📊 Attempting database initialization...")
    init_db()
    print("✅ Database initialized successfully")
except Exception as e:
    print(f"⚠️  Database init warning: {str(e)}")
    import traceback
    traceback.print_exc()

app = Flask(__name__)

# ==================== SECURITY HARDENING ====================
_rate_limit_lock = threading.Lock()
_rate_limit_attempts = {}

LOGIN_RATE_WINDOW_SECONDS = 15 * 60
LOGIN_RATE_MAX_ATTEMPTS = 10
RESET_RATE_WINDOW_SECONDS = 15 * 60
RESET_RATE_MAX_ATTEMPTS = 8


@app.template_filter('from_json')
def from_json_filter(value):
    """Safely parse JSON strings in templates."""
    if value is None:
        return []
    if isinstance(value, (list, dict)):
        return value
    try:
        parsed = json.loads(value)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return []

# Check secret key
secret_key = os.getenv('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')
if secret_key == 'dev-secret-key-change-in-production':
    print("⚠️  Using default secret key (not recommended for production)")
else:
    print("✅ Custom secret key configured")
app.secret_key = secret_key

app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = os.getenv('SESSION_COOKIE_SECURE', '0') == '1'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=12)

# Enable compression for all responses
Compress(app)

# Configuration
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create upload folder if it doesn't exist
try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    os.makedirs('gym_data', exist_ok=True)
    os.makedirs('static/uploads', exist_ok=True)
    print(f"✅ Directories created: {UPLOAD_FOLDER}, gym_data, static/uploads")
except Exception as e:
    print(f"⚠️ Directory creation warning: {str(e)}")

# Ensure users.json exists
if not os.path.exists('users.json'):
    print("Creating users.json...")
    with open('users.json', 'w') as f:
        json.dump({}, f)

auth_manager = AuthManager()
email_sender = EmailSender()
payment_manager = PaymentManager()

# Initialize Security Manager
from security_manager import SecurityManager
from models import get_session
security_manager = SecurityManager(get_session)

# Initialize modular routes (Required for Gunicorn production)
init_upgrade_routes(app, auth_manager, security_manager)

# ==================== BACKGROUND SCHEDULER (Phase 2) ====================

def run_scheduled_summaries():
    """Background task to send daily summaries to all gym owners"""
    with app.app_context():
        session = get_session()
        try:
            gyms = session.query(Gym).all()
            for gym in gyms:
                auto_man = AutomationManager(session, email_sender)
                auto_man.generate_daily_business_summary(gym.id)
        except Exception as e:
            print(f"❌ Scheduler Error (Summaries): {str(e)}")
        finally:
            session.close()

def run_scheduled_automations():
    """Background task to run daily reminders and wishes for all gyms"""
    with app.app_context():
        session = get_session()
        try:
            gyms = session.query(Gym).all()
            for gym in gyms:
                auto_man = AutomationManager(session, email_sender)
                auto_man.run_daily_automations(gym.id)
        except Exception as e:
            print(f"❌ Scheduler Error (Automations): {str(e)}")
        finally:
            session.close()

# Start the scheduler
if BackgroundScheduler is not None:
    try:
        scheduler = BackgroundScheduler(daemon=True)
        scheduler.add_job(func=run_scheduled_summaries, trigger="cron", hour=23, minute=59)
        scheduler.add_job(func=run_scheduled_automations, trigger="cron", hour=9, minute=0)
        scheduler.start()
        print("⏰ Background Scheduler started successfully (Summaries @ 23:59, Automations @ 09:00)")
    except Exception as scheduler_start_error:
        scheduler = None
        print(f"⚠️ Scheduler initialization failed, continuing without background jobs: {scheduler_start_error}")
else:
    scheduler = None

def get_gym():
    """Get GymManager instance for logged-in user"""
    if 'logged_in' not in session:
        return None
    username = session.get('username')

    if auth_manager.legacy or not getattr(auth_manager, 'session', None):
        return GymManager(username)

    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        return None

    if (user.role or '').lower() == 'staff':
        access = auth_manager.session.query(StaffAccess).filter_by(
            staff_user_id=user.id,
            is_active=True
        ).order_by(StaffAccess.created_at.desc()).first()

        if access:
            owner = auth_manager.session.query(User).filter_by(id=access.owner_user_id).first()
            if owner:
                return GymManager(owner.email)
        return None

    return GymManager(username)


def _get_company_details():
    """Company profile values for public landing page."""
    return {
        'name': os.getenv('COMPANY_NAME', 'Gym Manager Pro'),
        'tagline': os.getenv('COMPANY_TAGLINE', 'Professional SaaS solutions for gyms and fitness studios'),
        'phone': os.getenv('COMPANY_PHONE', '+92-300-0000000'),
        'email': os.getenv('COMPANY_EMAIL', 'support@gymmanagerpro.com'),
        'address': os.getenv('COMPANY_ADDRESS', 'Main Boulevard, Lahore, Pakistan'),
        'website': os.getenv('COMPANY_WEBSITE', 'https://gymmanagerpro.com')
    }


def _get_month_floor(gym, default_year=1970):
    """Return the earliest selectable month start date for a gym (no fixed past limit)."""
    floor_date = datetime(default_year, 1, 1)
    if not gym or getattr(gym, 'legacy', False) or not getattr(gym, 'session', None) or not getattr(gym, 'gym', None):
        return floor_date

    earliest_dates = []

    try:
        first_member = (
            gym.session.query(Member)
            .filter(Member.gym_id == gym.gym.id, Member.joined_date.isnot(None))
            .order_by(Member.joined_date.asc())
            .first()
        )
        if first_member and first_member.joined_date:
            earliest_dates.append(datetime(first_member.joined_date.year, first_member.joined_date.month, 1))
    except Exception:
        pass

    try:
        first_fee = (
            gym.session.query(Fee)
            .join(Member, Fee.member_id == Member.id)
            .filter(Member.gym_id == gym.gym.id, Fee.month.isnot(None))
            .order_by(Fee.month.asc())
            .first()
        )
        if first_fee and first_fee.month:
            parsed = datetime.strptime(str(first_fee.month)[:7], '%Y-%m')
            earliest_dates.append(datetime(parsed.year, parsed.month, 1))
    except Exception:
        pass

    if earliest_dates:
        return min(min(earliest_dates), floor_date)
    return floor_date


def _build_available_months(gym=None, future_months=0, as_dict=False, descending=True):
    """Build month options from earliest known gym data up to now (+ optional future months)."""
    start_date = _get_month_floor(gym)
    current_date = datetime.now().replace(day=1)

    total_months = ((current_date.year - start_date.year) * 12 + (current_date.month - start_date.month) + 1) + max(0, future_months)
    month_list = []

    for offset in range(total_months):
        month_index = (start_date.month - 1) + offset
        year = start_date.year + (month_index // 12)
        month = (month_index % 12) + 1
        month_date = datetime(year, month, 1)

        if as_dict:
            month_list.append({
                'value': month_date.strftime('%Y-%m'),
                'label': month_date.strftime('%B %Y')
            })
        else:
            month_list.append(month_date.strftime('%Y-%m'))

    if descending:
        month_list.reverse()
    return month_list


def _build_legacy_backup_payload(gym):
    """Build backup payload in legacy JSON-compatible structure."""
    if gym.legacy:
        payload = gym.data.copy() if isinstance(gym.data, dict) else {}
    else:
        payload = {
            'members': {},
            'fees': {},
            'attendance': {},
            'expenses': [],
            'gym_details': gym.get_gym_details()
        }

        members = gym.session.query(Member).filter_by(gym_id=gym.gym.id).all() if gym.gym else []
        for member in members:
            member_id = str(member.id)
            payload['members'][member_id] = {
                'id': member_id,
                'name': member.name,
                'phone': member.phone,
                'photo': member.photo_url,
                'membership_type': member.membership_type,
                'joined_date': member.joined_date.strftime('%Y-%m-%d') if member.joined_date else None,
                'is_trial': bool(member.is_trial),
                'trial_end_date': member.trial_end_date.strftime('%Y-%m-%d') if member.trial_end_date else None,
                'is_active': bool(member.is_active),
                'email': member.email
            }

            fee_records = gym.session.query(Fee).filter_by(member_id=member.id).all()
            payload['fees'][member_id] = {}
            for fee in fee_records:
                payload['fees'][member_id][fee.month] = {
                    'amount': float(fee.amount),
                    'date': fee.paid_date.strftime('%Y-%m-%d %H:%M:%S') if fee.paid_date else datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                }

            attendance_records = gym.get_attendance(member.id)
            payload['attendance'][member_id] = attendance_records if attendance_records else []

        payload['expenses'] = gym.get_expenses()

    try:
        payload['backup_meta'] = {
            'generated_at': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
            'format': 'legacy_compatible',
            'owner': session.get('username', 'unknown')
        }
    except Exception:
        pass

    return payload


def _parse_excel_backup_to_legacy(file_obj):
    """Parse old-structure Excel backup into legacy-compatible JSON dict."""
    from openpyxl import load_workbook

    workbook = load_workbook(file_obj, data_only=True)

    def _sheet_rows(sheet_name):
        if sheet_name not in workbook.sheetnames:
            return []
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
        output = []
        for row in rows[1:]:
            if row is None:
                continue
            row_map = {}
            has_value = False
            for index, value in enumerate(row):
                key = headers[index] if index < len(headers) else f'col_{index}'
                if key:
                    row_map[key] = value
                if value not in (None, ''):
                    has_value = True
            if has_value:
                output.append(row_map)
        return output

    members_rows = _sheet_rows('members')
    fees_rows = _sheet_rows('fees')
    attendance_rows = _sheet_rows('attendance')
    expenses_rows = _sheet_rows('expenses')
    info_rows = _sheet_rows('backup_info')

    members = {}
    fees = {}
    attendance = {}
    expenses = []

    def _as_text(value, default=''):
        if value is None:
            return default
        if hasattr(value, 'strftime'):
            try:
                if ' ' in default:
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                return value.strftime('%Y-%m-%d')
            except Exception:
                pass
        return str(value).strip()

    def _first_non_empty(row, keys, default=None):
        for key in keys:
            value = row.get(key)
            if value not in (None, ''):
                return value
        return default

    def _as_bool(value, default=False):
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        text = str(value).strip().lower()
        if text in ['1', 'true', 'yes', 'y']:
            return True
        if text in ['0', 'false', 'no', 'n']:
            return False
        return default

    for idx, row in enumerate(members_rows, start=1):
        member_id = _as_text(_first_non_empty(row, ['id', 'member_id'], idx))
        phone_value = _first_non_empty(row, ['phone', 'mobile', 'contact', 'phone_number'], '')
        joined_value = _first_non_empty(row, ['joined_date', 'join_date', 'joined'], datetime.utcnow())
        members[member_id] = {
            'id': member_id,
            'name': _as_text(_first_non_empty(row, ['name', 'member_name'], ''), ''),
            'phone': _as_text(phone_value, ''),
            'email': _as_text(_first_non_empty(row, ['email', 'mail'], ''), '') or None,
            'photo': _as_text(_first_non_empty(row, ['photo', 'photo_url', 'image'], ''), '') or None,
            'membership_type': _as_text(_first_non_empty(row, ['membership_type', 'type', 'plan'], 'Gym'), 'Gym'),
            'joined_date': _as_text(joined_value, datetime.utcnow().strftime('%Y-%m-%d')),
            'is_trial': _as_bool(_first_non_empty(row, ['is_trial', 'trial'], False), False),
            'trial_end_date': _as_text(_first_non_empty(row, ['trial_end_date', 'trial_end'], ''), '') or None,
            'is_active': _as_bool(row.get('is_active'), True)
        }
        fees.setdefault(member_id, {})
        attendance.setdefault(member_id, [])

    for row in fees_rows:
        member_id = _as_text(_first_non_empty(row, ['member_id', 'id'], ''), '')
        month = _as_text(_first_non_empty(row, ['month', 'paid_month'], ''), '')
        if not member_id or not month:
            continue
        fees.setdefault(member_id, {})
        fees[member_id][month] = {
            'amount': float(_first_non_empty(row, ['amount', 'fee', 'paid_amount'], 0) or 0),
            'date': _as_text(_first_non_empty(row, ['date', 'paid_date', 'timestamp'], datetime.utcnow()), datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'))
        }

    for row in attendance_rows:
        member_id = _as_text(_first_non_empty(row, ['member_id', 'id'], ''), '')
        if not member_id:
            continue
        attendance.setdefault(member_id, [])
        confidence_value = _first_non_empty(row, ['confidence', 'score'], None)
        try:
            confidence_value = float(confidence_value) if confidence_value not in (None, '') else None
        except Exception:
            confidence_value = None
        attendance[member_id].append({
            'timestamp': _as_text(_first_non_empty(row, ['timestamp', 'check_in_time', 'date'], datetime.utcnow()), datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')),
            'emotion': _as_text(_first_non_empty(row, ['emotion', 'mood'], ''), '') or None,
            'confidence': confidence_value
        })

    for row in expenses_rows:
        expenses.append({
            'date': _as_text(_first_non_empty(row, ['date', 'expense_date'], ''), ''),
            'category': _as_text(_first_non_empty(row, ['category', 'type'], ''), ''),
            'amount': float(_first_non_empty(row, ['amount', 'value'], 0) or 0),
            'description': _as_text(_first_non_empty(row, ['description', 'note', 'notes'], ''), '')
        })

    gym_details = {'name': 'Gym Manager', 'logo': None, 'currency': '$'}
    for row in info_rows:
        key = str(row.get('key') or '').strip().lower()
        value = row.get('value')
        if key == 'gym_name' and value:
            gym_details['name'] = str(value)
        elif key == 'currency' and value:
            gym_details['currency'] = str(value)

    return {
        'members': members,
        'fees': fees,
        'attendance': attendance,
        'expenses': expenses,
        'gym_details': gym_details
    }


FEATURE_KEYS = [
    'basic_dashboard',
    'member_management',
    'payment_tracking',
    'basic_reports',
    'mobile_app',
    'email_support',
    'marketing_automation',
    'advanced_analytics',
    'ai_churn_prediction',
    'multi_gym_management',
    'priority_support',
    'custom_branding',
    'bulk_operations',
    'advanced_reports',
    'api_access',
    'webhooks',
    'white_label',
    'custom_domain',
    'dedicated_account_manager',
    'sla_guarantee',
    '24_7_support',
    'on_premise_deployment',
    'custom_development',
    'phone_support',
    'training_sessions',
    'data_migration',
    'compliance_support',
    'multi_region_deployment'
]


def _get_current_user():
    """Get current DB user record for logged-in session, if available."""
    if 'logged_in' not in session:
        return None
    username = session.get('username')
    if not username or auth_manager.legacy:
        return None
    return auth_manager.session.query(User).filter_by(email=username).first()


def _is_subscription_active_user(user):
    """Check whether subscription is active for a DB user."""
    if not user:
        return False
    try:
        return auth_manager.is_subscription_active(user.email)
    except Exception:
        return False


def _has_feature_access(user, feature_key):
    """Feature check with active-subscription enforcement."""
    if auth_manager.legacy:
        return True
    if not user:
        return False
    if getattr(user, 'market', None) == 'VIP':
        return True
    if getattr(user, 'subscription_status', None) == 'trial':
        return _is_subscription_active_user(user)
    if not _is_subscription_active_user(user):
        return False
    return TierManager.has_feature(user, feature_key)


def _detect_payment_market(user=None):
    """Resolve effective payment market from user profile first, then request hints."""
    if user and getattr(user, 'market', None) in ['PK', 'US']:
        return user.market

    country_headers = [
        'CF-IPCountry',
        'CloudFront-Viewer-Country',
        'X-Country-Code',
        'X-Country',
        'X-App-Country'
    ]
    for header in country_headers:
        country = request.headers.get(header, '').strip().upper()
        if country == 'PK':
            return 'PK'
        if country == 'US':
            return 'US'

    accept_language = request.headers.get('Accept-Language', '').upper()
    if '-PK' in accept_language or '_PK' in accept_language:
        return 'PK'

    return 'US'


def _require_feature(feature_key, upgrade_message):
    """Return redirect response when feature is not available, else None."""
    if auth_manager.legacy:
        return None
    user = _get_current_user()
    if not user:
        return redirect(url_for('auth'))
    if not _has_feature_access(user, feature_key):
        flash(upgrade_message, 'warning')
        return redirect(url_for('subscription_plans'))
    return None


def _build_feature_paths():
    """Return canonical navigation paths for feature quick-actions."""
    def _safe_url(endpoint):
        try:
            return url_for(endpoint)
        except Exception:
            return '#'

    return {
        'member_management': _safe_url('add_member'),
        'payment_tracking': _safe_url('fees'),
        'basic_reports': _safe_url('reports'),
        'mobile_app': _safe_url('scanner'),
        'advanced_analytics': _safe_url('advanced_analytics'),
        'bulk_operations': _safe_url('bulk_operations'),
        'advanced_reports': _safe_url('export_center'),
        'webhooks': _safe_url('webhooks'),
        'api_access': _safe_url('webhooks'),
        'marketing_automation': _safe_url('schedule'),
        'multi_gym_management': _safe_url('super_admin'),
        'priority_support': _safe_url('subscription_plans'),
        'custom_branding': _safe_url('settings'),
        'white_label': _safe_url('settings'),
        'custom_domain': _safe_url('settings'),
        'dedicated_account_manager': _safe_url('subscription_plans'),
        'sla_guarantee': _safe_url('subscription_plans'),
        '24_7_support': _safe_url('subscription_plans'),
        'on_premise_deployment': _safe_url('subscription_plans'),
        'custom_development': _safe_url('subscription_plans'),
        'phone_support': _safe_url('subscription_plans'),
        'training_sessions': _safe_url('subscription_plans'),
        'data_migration': _safe_url('bulk_import'),
        'compliance_support': _safe_url('subscription_plans'),
        'multi_region_deployment': _safe_url('subscription_plans')
    }

@app.context_processor
def inject_gym_details():
    context = {}
    def _market_currency_symbol(market):
        return 'Rs' if market == 'PK' else '$'

    gym = get_gym()
    if gym:
        details = gym.get_gym_details()
        if 'currency' not in details: details['currency'] = '$'
        context['gym_details'] = details
    else:
        context['gym_details'] = {'name': 'Gym Manager', 'logo': None, 'currency': '$'}
        
    # Initialize with no access for anonymous users
    context['user_plan'] = 'starter'
    context['subscription_active'] = False
    context['feature_access'] = {key: False for key in FEATURE_KEYS}
    context['feature_paths'] = _build_feature_paths()
    context['payment_market'] = _detect_payment_market()

    if 'logged_in' in session:
        username = session.get('username')
        if auth_manager.legacy:
            # Legacy JSON mode: full access
            user_data = auth_manager.users.get(username, {})
            context['user_plan'] = user_data.get('plan', 'standard')
            context['subscription_active'] = True
            context['feature_access'] = {key: True for key in FEATURE_KEYS}
        else:
            # Database mode: check user subscription status
            user = auth_manager.session.query(User).filter_by(email=username).first()
            if user:
                # Check if user is VIP (lifetime access)
                if getattr(user, 'market', None) == 'VIP':
                    context['user_plan'] = 'vip'
                    context['subscription_active'] = True
                    # VIP users get ALL features enabled
                    context['feature_access'] = {key: True for key in FEATURE_KEYS}
                else:
                    # Regular users: check subscription status
                    context['user_plan'] = getattr(user, 'subscription_tier', 'starter') or 'starter'
                    context['subscription_active'] = _is_subscription_active_user(user)
                    # Build feature access based on subscription
                    context['feature_access'] = {
                        key: _has_feature_access(user, key) for key in FEATURE_KEYS
                    }
                context['payment_market'] = _detect_payment_market(user)
            else:
                context['user_plan'] = 'starter'
                context['subscription_active'] = False

    # Currency display according to detected region, unless user set a custom currency symbol.
    default_like_symbols = {'$', 'USD', 'Rs', 'PKR', None, ''}
    current_currency = context.get('gym_details', {}).get('currency')
    if current_currency in default_like_symbols:
        context['gym_details']['currency'] = _market_currency_symbol(context.get('payment_market'))
    
    return context

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _client_ip():
    forwarded = request.headers.get('X-Forwarded-For', '').strip()
    if forwarded:
        return forwarded.split(',')[0].strip()
    return request.remote_addr or 'unknown'


def _is_rate_limited(key, max_attempts, window_seconds):
    now = datetime.utcnow()
    window_start = now - timedelta(seconds=window_seconds)
    with _rate_limit_lock:
        attempts = _rate_limit_attempts.get(key, [])
        attempts = [t for t in attempts if t >= window_start]
        _rate_limit_attempts[key] = attempts
        return len(attempts) >= max_attempts


def _register_rate_limit_attempt(key):
    now = datetime.utcnow()
    with _rate_limit_lock:
        attempts = _rate_limit_attempts.get(key, [])
        attempts.append(now)
        _rate_limit_attempts[key] = attempts[-100:]


def _clear_rate_limit_attempts(key):
    with _rate_limit_lock:
        if key in _rate_limit_attempts:
            del _rate_limit_attempts[key]

# STRIPE CONFIGURATION - Loaded from environment variables
# Get your keys from https://dashboard.stripe.com/apikeys
app.config['STRIPE_PUBLIC_KEY'] = os.getenv('STRIPE_PUBLIC_KEY', '')
app.config['STRIPE_SECRET_KEY'] = os.getenv('STRIPE_SECRET_KEY', '')
stripe.api_key = app.config['STRIPE_SECRET_KEY']

# GOOGLE OAUTH CONFIGURATION - Loaded from environment variables
app.config['GOOGLE_CLIENT_ID'] = os.getenv('GOOGLE_CLIENT_ID', '')

# JAZZCASH CONFIGURATION
app.config['JAZZCASH_MERCHANT_ID'] = os.getenv('JAZZCASH_MERCHANT_ID', '')
app.config['JAZZCASH_PASSWORD'] = os.getenv('JAZZCASH_PASSWORD', '')
app.config['JAZZCASH_INTEGRITY_SALT'] = os.getenv('JAZZCASH_INTEGRITY_SALT', '')
app.config['JAZZCASH_RETURN_URL'] = os.getenv('JAZZCASH_RETURN_URL', 'http://localhost:5000/jazzcash_return')


@app.before_request
def check_subscription():
    endpoint = request.endpoint

    # Ignore unknown endpoints (CLI/debug/static edge-cases)
    if not endpoint:
        return

    # Always-public endpoints
    public_endpoints = {
        'index', 'auth', 'google_login', 'logout', 'forgot_password', 'reset_password',
        'static', 'subscription', 'subscription_plans', 'upgrade_tier', 'upgrade_success',
        'activate_trial', 'create_checkout_session', 'payment_success', 'payment_cancel',
        'stripe_webhook', 'initiate_payment', 'payment_callback', 'fix_database_schema', 'healthz',
        'super_admin', 'approve_payment'
    }

    if endpoint in public_endpoints:
        return

    # Any non-public endpoint requires login
    if not session.get('logged_in'):
        return redirect(url_for('auth'))

    # Legacy mode keeps full access (backward compatibility)
    if auth_manager.legacy:
        return

    user = _get_current_user()
    if not user:
        session.clear()
        flash('Session expired. Please login again.', 'warning')
        return redirect(url_for('auth'))

    endpoint_feature_map = {
        # Core starter access
        'dashboard': 'basic_dashboard',
        'analytics': 'basic_reports',
        'reports': 'basic_reports',
        'schedule': 'basic_dashboard',
        'book_class': 'basic_dashboard',
        'expenses': 'basic_reports',
        'delete_expense': 'basic_reports',
        'settings': 'basic_dashboard',
        'restore_backup': 'basic_dashboard',
        'merge_duplicates': 'basic_dashboard',
        'reset_admin': 'basic_dashboard',

        # Member management
        'add_member': 'member_management',
        'member_details': 'member_management',
        'add_member_note': 'member_management',
        'delete_member_note': 'member_management',
        'add_measurement': 'member_management',
        'delete_measurement': 'member_management',
        'edit_member': 'member_management',
        'delete_member': 'member_management',
        'bulk_import': 'member_management',
        'import_preview': 'member_management',
        'download_template': 'member_management',

        # Payment tracking
        'fees': 'payment_tracking',
        'edit_fee_record': 'payment_tracking',
        'delete_fee_record': 'payment_tracking',
        'generate_receipt': 'payment_tracking',

        # Mobile app/PWA access points
        'scanner': 'mobile_app',
        'scan_check': 'mobile_app',
        'generate_wallet_pass': 'mobile_app',

        # Pro+ features
        'advanced_analytics': 'advanced_analytics',
        'bulk_operations': 'bulk_operations',
        'bulk_payment': 'bulk_operations',

        # Advanced reports (Pro+)
        'download_excel': 'advanced_reports',
        'export_center': 'advanced_reports',
        'export_members_complete': 'advanced_reports',
        'export_revenue_report': 'advanced_reports',
        'export_unpaid_members': 'advanced_reports',

        # Enterprise+
        'webhooks': 'webhooks',
        'create_webhook': 'webhooks',
        'toggle_webhook': 'webhooks',
        'delete_webhook': 'webhooks',
        'test_webhook': 'webhooks',
        'webhook_logs': 'webhooks',
        'chatbot_api': 'api_access'
    }

    feature_messages = {
        'basic_dashboard': 'Please activate your subscription to continue.',
        'member_management': 'Member Management is not available in your current plan.',
        'payment_tracking': 'Payment Tracking is not available in your current plan.',
        'basic_reports': 'Basic Reports are not available in your current plan.',
        'mobile_app': 'Mobile App (PWA) is not available in your current plan.',
        'advanced_analytics': 'Advanced Analytics is available in Professional tier and above.',
        'bulk_operations': 'Bulk Operations are available in Professional tier and above.',
        'advanced_reports': 'Advanced report exports are available in Professional tier and above.',
        'webhooks': 'Webhooks are only available in Enterprise tier and above.',
        'api_access': 'API Access is only available in Enterprise tier and above.'
    }

    required_feature = endpoint_feature_map.get(endpoint)
    if required_feature and not _has_feature_access(user, required_feature):
        flash(feature_messages.get(required_feature, 'This feature is not available in your current plan.'), 'warning')
        return redirect(url_for('subscription_plans'))


@app.after_request
def add_no_cache_headers(response):
    """Add security and cache headers."""
    try:
        response.headers.setdefault('X-Content-Type-Options', 'nosniff')
        response.headers.setdefault('X-Frame-Options', 'SAMEORIGIN')
        response.headers.setdefault('Referrer-Policy', 'strict-origin-when-cross-origin')
        response.headers.setdefault('Permissions-Policy', 'camera=(), microphone=(), geolocation=()')

        if response.mimetype == 'text/html':
            response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
    except Exception:
        pass
    return response

# Subscription and Tier routes are now handled by tier_routes.py


#  ========== WEBHOOK MANAGEMENT ROUTES ==========

@app.route('/webhooks')
def webhooks():
    """Webhook management page"""
    from webhook_manager import WebhookManager, Webhook
    
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))
    
    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        return redirect(url_for('auth'))
    
    gate = _require_feature('webhooks', 'Webhooks are only available in Enterprise tier and above')
    if gate:
        return gate
    
    # Get user's webhooks
    webhooks_list = WebhookManager.get_webhooks(user.id)
    
    return render_template('webhooks.html', 
                         webhooks=webhooks_list,
                         available_events=WebhookManager.EVENTS)


@app.route('/webhooks/create', methods=['POST'])
def create_webhook():
    """Create new webhook"""
    from webhook_manager import Webhook
    from models import get_session
    import secrets
    
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))
    
    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        return redirect(url_for('auth'))

    gate = _require_feature('webhooks', 'Webhooks are only available in Enterprise tier and above')
    if gate:
        return gate
    
    name = request.form.get('name')
    url = request.form.get('url')
    events = request.form.getlist('events')
    secret = request.form.get('secret') or secrets.token_urlsafe(32)
    
    # Create webhook
    webhook_session = get_session()
    webhook = Webhook(
        user_id=user.id,
        name=name,
        url=url,
        events=json.dumps(events),
        secret=secret
    )
    
    webhook_session.add(webhook)
    webhook_session.commit()
    webhook_session.close()
    
    flash(f'✅ Webhook "{name}" created successfully!', 'success')
    return redirect(url_for('webhooks'))


@app.route('/webhooks/<int:webhook_id>/toggle', methods=['POST'])
def toggle_webhook(webhook_id):
    """Enable/disable webhook"""
    from webhook_manager import Webhook
    from models import get_session
    
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))

    gate = _require_feature('webhooks', 'Webhooks are only available in Enterprise tier and above')
    if gate:
        return gate

    webhook_session = get_session()
    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        webhook_session.close()
        return redirect(url_for('auth'))

    webhook = webhook_session.query(Webhook).get(webhook_id)
    
    if webhook and webhook.user_id == user.id:
        webhook.is_active = not webhook.is_active
        webhook_session.commit()
        
        status = "enabled" if webhook.is_active else "disabled"
        flash(f'Webhook {status}', 'success')
    
    webhook_session.close()
    return redirect(url_for('webhooks'))


@app.route('/webhooks/<int:webhook_id>/delete', methods=['POST'])
def delete_webhook(webhook_id):
    """Delete webhook"""
    from webhook_manager import Webhook
    from models import get_session
    
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))

    gate = _require_feature('webhooks', 'Webhooks are only available in Enterprise tier and above')
    if gate:
        return gate

    webhook_session = get_session()
    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        webhook_session.close()
        return redirect(url_for('auth'))

    webhook = webhook_session.query(Webhook).get(webhook_id)
    
    if webhook and webhook.user_id == user.id:
        webhook_session.delete(webhook)
        webhook_session.commit()
        flash('🗑️ Webhook deleted', 'success')
    
    webhook_session.close()
    return redirect(url_for('webhooks'))


@app.route('/webhooks/<int:webhook_id>/test', methods=['POST'])
def test_webhook(webhook_id):
    """Test webhook with sample data"""
    from webhook_manager import WebhookManager
    from webhook_manager import Webhook
    from models import get_session
    
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))

    gate = _require_feature('webhooks', 'Webhooks are only available in Enterprise tier and above')
    if gate:
        return gate
    
    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        return redirect(url_for('auth'))

    webhook_session = get_session()
    webhook = webhook_session.query(Webhook).get(webhook_id)
    if not webhook or webhook.user_id != user.id:
        webhook_session.close()
        flash('Webhook not found.', 'error')
        return redirect(url_for('webhooks'))
    webhook_session.close()
    
    # Trigger test event
    WebhookManager.trigger_event(
        user_id=user.id,
        event_type='member.created',
        data={
            'test': True,
            'member_id': 999,
            'name': 'Test Member',
            'message': 'This is a test webhook delivery'
        }
    )
    
    flash('🧪 Test webhook sent! Check logs for results.', 'info')
    return redirect(url_for('webhooks'))


@app.route('/webhooks/<int:webhook_id>/logs')
def webhook_logs(webhook_id):
    """View webhook logs"""
    from webhook_manager import WebhookManager
    from webhook_manager import Webhook
    from models import get_session
    
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))

    gate = _require_feature('webhooks', 'Webhooks are only available in Enterprise tier and above')
    if gate:
        return gate

    user = auth_manager.session.query(User).filter_by(email=username).first()
    if not user:
        return redirect(url_for('auth'))

    webhook_session = get_session()
    webhook = webhook_session.query(Webhook).get(webhook_id)
    if not webhook or webhook.user_id != user.id:
        webhook_session.close()
        flash('Webhook not found.', 'error')
        return redirect(url_for('webhooks'))
    webhook_session.close()
    
    logs = WebhookManager.get_webhook_logs(webhook_id, limit=100)
    
    return render_template('webhook_logs.html', logs=logs, webhook_id=webhook_id)

@app.route('/create_checkout_session', methods=['POST'])
def create_checkout_session():
    username = session.get('username')
    try:
        checkout_session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{
                'price_data': {
                    'currency': 'usd',
                    'product_data': {
                        'name': 'Gym Manager Pro Subscription',
                        'images': ['https://i.imgur.com/EHyR2nP.png'],
                    },
                    'unit_amount': 6000, # $60.00
                },
                'quantity': 1,
            }],
            mode='payment',
            success_url=url_for('payment_success', _external=True) + '?session_id={CHECKOUT_SESSION_ID}',
            cancel_url=url_for('payment_cancel', _external=True),
            client_reference_id=username,
        )
        return redirect(checkout_session.url, code=303)
    except Exception as e:
        flash(f'Error creating payment session: {str(e)}', 'error')
        return redirect(url_for('subscription'))


@app.route('/create_billing_portal_session', methods=['POST'])
def create_billing_portal_session():
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))

    if not app.config.get('STRIPE_SECRET_KEY'):
        flash('Billing portal is not configured yet. Please set Stripe keys in environment settings.', 'warning')
        return redirect(url_for('settings'))

    try:
        customers = stripe.Customer.list(email=username, limit=1)
        customer_id = None

        if customers and getattr(customers, 'data', None):
            customer_id = customers.data[0].id
        else:
            customer = stripe.Customer.create(email=username, metadata={'source': 'gym-manager-saas'})
            customer_id = customer.id

        portal_session = stripe.billing_portal.Session.create(
            customer=customer_id,
            return_url=url_for('settings', _external=True)
        )
        return redirect(portal_session.url, code=303)
    except Exception as e:
        flash(f'Unable to open billing portal: {str(e)}', 'error')
        return redirect(url_for('settings'))

@app.route('/payment_success')
def payment_success():
    username = session.get('username')
    if not username:
        return redirect(url_for('auth'))

    session_id = request.args.get('session_id')
    if not session_id:
        flash('Payment verification failed: missing session ID.', 'error')
        return redirect(url_for('subscription'))

    try:
        checkout_session = stripe.checkout.Session.retrieve(session_id)

        if checkout_session.payment_status != 'paid':
            flash(f'Payment not completed (status: {checkout_session.payment_status}).', 'error')
            return redirect(url_for('subscription'))

        session_user = checkout_session.client_reference_id or checkout_session.customer_email
        if session_user and session_user.lower() != username.lower():
            flash('Payment verification failed: account mismatch.', 'error')
            return redirect(url_for('subscription'))

        metadata = getattr(checkout_session, 'metadata', {}) or {}
        metadata_days = metadata.get('subscription_days', '30')
        try:
            subscription_days = int(metadata_days)
        except Exception:
            subscription_days = int(session.pop('pending_subscription_days', 30) or 30)

        auth_manager.extend_subscription(username, days=subscription_days)
        auth_manager.set_market(username, 'US')
        user = auth_manager.session.query(User).filter_by(email=username).first()
        if user:
            user.subscription_status = 'active'
            user.subscription_tier = metadata.get('subscription_tier', 'enterprise_plus')
            user.billing_cycle = 'monthly'
            auth_manager.session.commit()
        flash('Payment Successful! Thank you for your subscription. ✅', 'success')
        session.pop('needs_payment', None)
        session.pop('pending_subscription_days', None)
        return redirect(url_for('dashboard'))
    except Exception as e:
        flash(f'Payment verification failed: {str(e)}', 'error')
        return redirect(url_for('subscription'))

@app.route('/payment_cancel')
def payment_cancel():
    flash('Payment cancelled.', 'info')
    return redirect(url_for('subscription'))


@app.route('/stripe/webhook', methods=['POST'])
def stripe_webhook():
    payload = request.data
    sig_header = request.headers.get('Stripe-Signature')
    webhook_secret = os.getenv('STRIPE_WEBHOOK_SECRET', '')

    if not webhook_secret:
        print('⚠️ Stripe webhook received but STRIPE_WEBHOOK_SECRET is not configured; ignoring event')
        return jsonify({'status': 'ignored', 'reason': 'webhook secret not configured'}), 200

    try:
        event = stripe.Webhook.construct_event(payload, sig_header, webhook_secret)
    except ValueError:
        return 'Invalid payload', 400
    except stripe.error.SignatureVerificationError:
        return 'Invalid signature', 400

    if event['type'] == 'checkout.session.completed':
        checkout_session = event['data']['object']

        if checkout_session.get('payment_status') == 'paid':
            target_user = checkout_session.get('client_reference_id') or checkout_session.get('customer_email')
            if target_user:
                metadata = checkout_session.get('metadata') or {}
                try:
                    subscription_days = int(metadata.get('subscription_days', 30))
                except Exception:
                    subscription_days = 30
                auth_manager.extend_subscription(target_user, days=subscription_days)
                auth_manager.set_market(target_user, 'US')
                user = auth_manager.session.query(User).filter_by(email=target_user).first()
                if user:
                    user.subscription_status = 'active'
                    user.subscription_tier = metadata.get('subscription_tier', 'enterprise_plus')
                    user.billing_cycle = 'monthly'
                    auth_manager.session.commit()

    return jsonify({'status': 'ok'}), 200


@app.route('/healthz')
def healthz():
    """Lightweight health endpoint for load balancers and uptime monitors."""
    return jsonify({
        'status': 'ok',
        'service': 'gym-manager',
        'mode': 'database' if not auth_manager.legacy else 'legacy',
        'timestamp': datetime.utcnow().isoformat() + 'Z'
    }), 200

# Trial activation is now handled by tier_routes.py

@app.route('/initiate_payment', methods=['POST'])
def initiate_payment():
    """Initiate payment with selected provider"""
    username = session.get('username')
    if not username:
        flash('Please login first', 'error')
        return redirect(url_for('auth'))
    
    provider = request.form.get('provider')  # 'jazzcash', 'easypaisa', 'stripe'
    valid_providers = {'jazzcash', 'easypaisa', 'stripe'}
    if provider not in valid_providers:
        flash('Invalid payment provider selected.', 'error')
        return redirect(url_for('subscription'))

    user = auth_manager.session.query(User).filter_by(email=username).first() if not auth_manager.legacy else None
    if not user:
        flash('User not found. Please login again.', 'error')
        return redirect(url_for('auth'))
    
    # Use actual email from database, not session username
    user_email = user.email
    payment_market = _detect_payment_market(user)

    if provider in ['jazzcash', 'easypaisa'] and payment_market != 'PK':
        flash('JazzCash and EasyPaisa are only available in Pakistan.', 'warning')
        return redirect(url_for('subscription'))

    plan = request.form.get('plan', '').strip().lower()
    amount_pkr = 5000  # Rs 5000 for PK
    amount_usd = 60    # $60 for International default
    subscription_days = 30
    subscription_tier = 'enterprise_plus'

    if provider == 'stripe' and plan == 'mini_monthly_us':
        amount_usd = 6
        subscription_days = 30
    
    # Determine amount based on provider
    if provider in ['jazzcash', 'easypaisa']:
        amount = amount_pkr
        currency = 'PKR'
    else:
        amount = amount_usd
        currency = 'USD'
    
    return_url = url_for('payment_callback', provider=provider, _external=True)
    
    # Initiate payment through PaymentManager
    session['pending_subscription_days'] = subscription_days

    result = payment_manager.initiate_payment(
        provider=provider,
        amount=amount,
        user_email=user_email,
        return_url=return_url,
        success_url=url_for('payment_success', _external=True),
        cancel_url=url_for('payment_cancel', _external=True),
        metadata={
            'subscription_days': str(subscription_days),
            'subscription_plan': plan or ('monthly_pk' if provider in ['jazzcash', 'easypaisa'] else 'monthly_us'),
            'subscription_tier': subscription_tier
        }
    )
    
    if not result.get('success'):
        flash(f'Payment initiation failed: {result.get("error")}', 'error')
        return redirect(url_for('subscription'))
    
    # For Stripe, redirect to checkout URL
    if provider == 'stripe':
        return redirect(result['session_url'])
    
    # For JazzCash/EasyPaisa, render auto-submit form
    return render_template('payment_redirect.html', 
                          post_url=result['post_url'],
                          form_data=result['form_data'])

@app.route('/payment_callback/<provider>', methods=['GET', 'POST'])
def payment_callback(provider):
    """Handle payment gateway callback"""
    username = session.get('username')
    if not username:
        flash('Session expired', 'error')
        return redirect(url_for('auth'))
    
    # Get response data
    if request.method == 'POST':
        response_data = request.form.to_dict()
    else:
        response_data = request.args.to_dict()
    
    # Verify payment
    success, message = payment_manager.verify_payment(provider, response_data)
    
    if success:
        subscription_days = int(session.pop('pending_subscription_days', 30) or 30)
        # Activate subscription
        auth_manager.extend_subscription(username, days=subscription_days)
        user = auth_manager.session.query(User).filter_by(email=username).first()
        if user:
            user.subscription_status = 'active'
            user.subscription_tier = 'enterprise_plus'
            user.billing_cycle = 'monthly'
            auth_manager.session.commit()
        
        # Set market based on provider
        if provider in ['jazzcash', 'easypaisa']:
            auth_manager.set_market(username, 'PK')
        else:
            auth_manager.set_market(username, 'US')
        
        flash(f'✅ Payment Successful! {message}', 'success')
        session.pop('needs_payment', None)
        return redirect(url_for('dashboard'))
    else:
        flash(f'❌ Payment Failed: {message}', 'error')
        return redirect(url_for('subscription'))
    
    if request.method == 'POST':
        if 'payment_proof' in request.files:
            file = request.files['payment_proof']
            if file and file.filename:
                filename = secure_filename(f"proof_{username}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                
                # Update user status to pending
                auth_manager.set_payment_pending(username, filename)
                
                flash('Proof uploaded! Waiting for admin approval.', 'success')
                return redirect(url_for('subscription'))
                
    return render_template('payment_manual.html')

# Admin Access Control - Loaded from environment variables
ADMIN_EMAILS = os.getenv('ADMIN_EMAILS', 'admin@gym.com').split(',')

@app.route('/super_admin')
def super_admin():
    if session.get('username') not in ADMIN_EMAILS:
        flash('Access Denied: Super Admin only.', 'error')
        return redirect(url_for('dashboard'))
    
    pending_users = auth_manager.get_pending_approvals()
    return render_template('super_admin.html', pending_users=pending_users)

@app.route('/approve_payment/<target_username>')
def approve_payment(target_username):
    if session.get('username') not in ADMIN_EMAILS:
        flash('Access Denied.', 'error')
        return redirect(url_for('dashboard'))
        
    # Verify admin here logic
    if auth_manager.approve_manual_payment(target_username):
        flash(f'User {target_username} approved!', 'success')
        
        # Optional: Add a real Stripe payment record or just the manual one (already done in AuthManager)
    else:
        flash('Approval failed.', 'error')
    return redirect(url_for('super_admin'))

@app.route('/')
def index():
    """Root route - dashboard for logged-in users, marketing landing for visitors."""
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))
    return render_template('landing.html', company_details=_get_company_details())

@app.route('/auth', methods=['GET', 'POST'])
def auth():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        action = request.form.get('action')
        referral_code = request.form.get('referral_code')
        
        if action == 'signup':
            if auth_manager.create_user(username, password, referral_code):
                flash('Account created successfully! Please login.', 'success')
                return redirect(url_for('auth'))
            else:
                if referral_code and auth_manager.apply_referral_code(username, referral_code):
                    flash('Referral code applied successfully. Your account now has unlimited access. Please login.', 'success')
                    return redirect(url_for('auth'))
                flash('Username already exists!', 'error')
        
        elif action == 'login':
            login_ip = _client_ip()
            login_user = (username or '').strip().lower()
            ip_key = f'login:ip:{login_ip}'
            user_key = f'login:user:{login_user}'

            if _is_rate_limited(ip_key, LOGIN_RATE_MAX_ATTEMPTS, LOGIN_RATE_WINDOW_SECONDS) or \
               _is_rate_limited(user_key, LOGIN_RATE_MAX_ATTEMPTS, LOGIN_RATE_WINDOW_SECONDS):
                flash('Too many login attempts. Please wait a few minutes and try again.', 'error')
                return redirect(url_for('auth'))

            if auth_manager.verify_user(username, password):
                session['logged_in'] = True
                session['username'] = username
                session.permanent = True
                _clear_rate_limit_attempts(ip_key)
                _clear_rate_limit_attempts(user_key)
                
                # Production Security: Log Audit & Session
                user_id = auth_manager.get_user_id(username)
                if user_id:
                    token = security_manager.create_session_token()
                    session['session_token'] = token
                    security_manager.track_session(user_id, token, request.remote_addr, request.user_agent.string)
                    security_manager.log_action(user_id, 'LOGIN_SUCCESS', {'ip': request.remote_addr}, request.remote_addr, request.user_agent.string)
                
                flash('Login successful!', 'success')
                return redirect(url_for('dashboard'))
            else:
                _register_rate_limit_attempt(ip_key)
                _register_rate_limit_attempt(user_key)
                user_id = auth_manager.get_user_id(username)
                if user_id:
                    security_manager.log_action(user_id, 'LOGIN_FAILURE', {'ip': request.remote_addr}, request.remote_addr, request.user_agent.string)
                flash('Invalid credentials!', 'error')
    
    return render_template('auth.html')


@app.route('/apply_referral', methods=['POST'])
def apply_referral():
    """Apply referral code for currently logged-in user."""
    next_page = request.form.get('next') or request.args.get('next') or 'settings'
    if not session.get('logged_in'):
        return redirect(url_for('auth'))

    username = session.get('username')
    referral_code = request.form.get('referral_code', '').strip()

    if not referral_code:
        flash('Please enter a referral code.', 'warning')
        return redirect(url_for(next_page))

    if auth_manager.apply_referral_code(username, referral_code):
        # Invalidate any cached session data to force fresh user fetch
        auth_manager.session.expunge_all()
        flash('✅ Referral code applied! Unlimited access activated. All features unlocked!', 'success')
        return redirect(url_for('dashboard'))

    flash('❌ Invalid referral code. Please check and try again.', 'error')
    return redirect(url_for(next_page))

@app.route('/google_login', methods=['POST'])
def google_login():
    token = request.form.get('credential')
    try:
        # Specify the CLIENT_ID of the app that accesses the backend (from environment variables)
        client_id = os.getenv('GOOGLE_CLIENT_ID', '')
        if not client_id:
            flash('Google Login not configured. Please contact administrator.', 'error')
            return redirect(url_for('auth'))
        idinfo = id_token.verify_oauth2_token(token, google_requests.Request(), client_id)

        # ID token is valid. Get the user's Google Account ID from the decoded token.
        email = idinfo['email']

        # Ensure user exists in our system
        if not auth_manager.user_exists(email):
            # Auto-signup with "GOOGLE" referral/plan logic if needed, or default free?
            # User asked for: "referral mera jo hoga wo bs free ho single time use"
            # We'll treat Google signups as standard/free for now.
            auth_manager.create_user(email, "GOOGLE_AUTH_USER", referral_code="GOOGLE_SIGNUP")
            flash(f'Account created with Google! Welcome, {email}.', 'success')

        session['logged_in'] = True
        session['username'] = email

        if auth_manager.user_exists(email):
             flash(f'Logged in as {email}!', 'success')

        return redirect(url_for('dashboard'))
    except ValueError:
        # Invalid token
        flash('Google Login failed! Invalid token.', 'error')
        return redirect(url_for('auth'))

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully!', 'success')
    return redirect(url_for('auth'))

@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    """Request password reset code"""
    if request.method == 'POST':
        reset_ip = _client_ip()
        reset_key = f'reset:ip:{reset_ip}'
        if _is_rate_limited(reset_key, RESET_RATE_MAX_ATTEMPTS, RESET_RATE_WINDOW_SECONDS):
            flash('Too many reset attempts. Please wait a few minutes before trying again.', 'error')
            return redirect(url_for('forgot_password'))

        email = request.form.get('email', '').strip()
        
        if not email:
            _register_rate_limit_attempt(reset_key)
            flash('Please enter your email address.', 'error')
            return redirect(url_for('forgot_password'))
        
        # Check if user exists
        if not auth_manager.user_exists(email):
            _register_rate_limit_attempt(reset_key)
            flash('No account found with that email.', 'error')
            return redirect(url_for('forgot_password'))
        
        # Generate reset code
        reset_code = auth_manager.generate_reset_code(email)
        
        if reset_code:
            _register_rate_limit_attempt(reset_key)
            # Try to send email
            from email_utils import EmailSender
            email_sender = EmailSender()
            
            if email_sender.is_configured():
                email_sender.send_reset_code(email, reset_code, email)
                flash(f'Reset code sent to {email}! Check your email.', 'success')
            else:
                # Email not configured - show code on screen (dev mode)
                flash(f'⚠️ Email not configured. Your reset code is: {reset_code}', 'warning')
            
            # Redirect to reset page
            return redirect(url_for('reset_password', email=email))
        else:
            _register_rate_limit_attempt(reset_key)
            flash('Error generating reset code. Please try again.', 'error')
    
    return render_template('forgot_password.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    """Verify code and reset password"""
    email = request.args.get('email') or request.form.get('email')
    
    if not email:
        flash('Invalid reset link.', 'error')
        return redirect(url_for('forgot_password'))
    
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        # Validate inputs
        if not code or not new_password or not confirm_password:
            flash('Please fill all fields.', 'error')
            return render_template('reset_password.html', email=email)
        
        if new_password != confirm_password:
            flash('Passwords do not match!', 'error')
            return render_template('reset_password.html', email=email)
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters.', 'error')
            return render_template('reset_password.html', email=email)
        
        # Verify code
        if auth_manager.verify_reset_code(email, code):
            # Reset password
            if auth_manager.reset_password(email, new_password):
                # Send confirmation email
                from email_utils import EmailSender
                email_sender = EmailSender()
                if email_sender.is_configured():
                    email_sender.send_password_changed_notification(email, email)
                
                flash('✅ Password reset successful! Please login with your new password.', 'success')
                return redirect(url_for('auth'))
            else:
                flash('Error resetting password. Please try again.', 'error')
        else:
            flash('❌ Invalid or expired code. Please request a new one.', 'error')
            return redirect(url_for('forgot_password'))
    
    return render_template('reset_password.html', email=email)


@app.route('/schedule', methods=['GET', 'POST'])
def schedule():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        day = request.form.get('day')
        time = request.form.get('time')
        instructor = request.form.get('instructor')
        capacity = request.form.get('capacity')
        
        gym.add_class(name, day, time, instructor, capacity)
        flash('Class added successfully!', 'success')
        return redirect(url_for('schedule'))
        
    return render_template('schedule.html', classes=gym.get_classes(), members=gym.get_all_members())

@app.route('/expenses', methods=['GET', 'POST'])
def expenses():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if request.method == 'POST':
        category = request.form.get('category')
        amount = float(request.form.get('amount') or 0)
        date = request.form.get('date')
        description = request.form.get('description', '')
        
        if gym.add_expense(category, amount, date, description):
            flash(f'Expense of {amount} recorded successfully!', 'success')
        else:
            flash('Failed to add expense!', 'error')
        
        return redirect(url_for('expenses'))
    
    # Get current month
    current_month = datetime.now().strftime('%Y-%m')
    
    # Get expenses for current month
    expenses_list = gym.get_expenses(current_month)
    
    # Calculate P&L
    pl_data = gym.calculate_profit_loss(current_month)
    
    # Available months for dropdown (no fixed past limit)
    available_months = _build_available_months(gym=gym, as_dict=True)
    
    return render_template('expenses.html',
                         expenses=expenses_list,
                         pl_data=pl_data,
                         current_month=current_month,
                         available_months=available_months,
                         gym_details=gym.get_gym_details())

@app.route('/delete_expense/<expense_id>', methods=['POST'])
def delete_expense(expense_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if gym.delete_expense(expense_id):
        flash('Expense deleted successfully!', 'success')
    else:
        flash('Failed to delete expense!', 'error')
    
    return redirect(url_for('expenses'))

@app.route('/book_class/<class_id>', methods=['POST'])
def book_class(class_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member_id = request.form.get('member_id')
    if gym.book_class(member_id, class_id):
        flash('Booking confirmed!', 'success')
    else:
        flash('Booking failed (Full or invalid)', 'error')
        
    return redirect(url_for('schedule'))

@app.route('/reports')
def reports():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    # Calculate stats
    total_members = len(gym.get_all_members())
    
    # Current month revenue
    current_month = datetime.now().strftime('%Y-%m')
    status = gym.get_payment_status(current_month)
    monthly_revenue = sum(m.get('amount', 0) for m in status['paid'])
    
    # Total check-ins
    total_checkins = 0
    if gym.legacy and hasattr(gym, 'data') and 'attendance' in gym.data:
        for visits in gym.data['attendance'].values():
            total_checkins += len(visits)
    elif not gym.legacy:
        # Database mode - count attendance records
        try:
            from models import get_session, Attendance
            session = get_session()
            if session:
                total_checkins = session.query(Attendance).count()
                session.close()
        except:
            total_checkins = 0
    
    # Revenue trend (last 6 months)
    revenue_months = []
    revenue_data = []
    for i in range(5, -1, -1):
        month = (datetime.now().replace(day=1) - timedelta(days=30*i)).strftime('%Y-%m')
        revenue_months.append(month)
        month_status = gym.get_payment_status(month)
        revenue_data.append(sum(m.get('amount', 0) for m in month_status['paid']))
    
    return render_template('reports.html',
                         total_members=total_members,
                         monthly_revenue=monthly_revenue,
                         total_checkins=total_checkins,
                         paid_count=len(status['paid']),
                         unpaid_count=len(status['unpaid']),
                         revenue_months=revenue_months,
                         revenue_data=revenue_data)

@app.route('/reset_admin')
def reset_admin():
    gym = get_gym()
    if gym:
        gym.reset_data()
        flash('All your data has been reset!', 'success')
    # Use referer check or just redirect dashboard to force re-login check? 
    # Actually, reset_data keeps the file but empties content. User is still logged in.
    return redirect(url_for('dashboard'))

@app.route('/dashboard')
def dashboard():
    try:
        gym = get_gym()
        if not gym: return redirect(url_for('auth'))

        # Generate months for dropdown (no fixed past limit)
        available_months = _build_available_months(gym=gym, as_dict=False)
        
        # Check if month requested
        current_month = request.args.get('month')
        if not current_month:
            current_month = datetime.now().strftime('%Y-%m')
            
        # ====================== ULTIMATE PERFORMANCE FIX ======================
        # Fetch EVERYTHING in one go using optimized SQL queries
        stats, alerts, charts = gym.get_dashboard_stats()
        revenue_amount = stats.get('revenue', 0) or 0
        revenue_number = float(revenue_amount)
        revenue_display = f"{revenue_number:,.0f}" if revenue_number.is_integer() else f"{revenue_number:,.2f}"
        
        # Fallback for handling empty data/errors
        revenue_trend = charts.get('revenue_trend', [])
        max_revenue = max([r['revenue'] for r in revenue_trend]) if revenue_trend else 1
        
        return render_template('dashboard_enhanced.html',
                            # Stats
                            total_members=stats.get('total_members', 0),
                            paid=stats.get('paid_list', []), # Fixed: Pass LIST not COUNT
                            unpaid=stats.get('unpaid_list', []), # Fixed: Pass LIST not COUNT
                            revenue=revenue_amount,
                            revenue_display=revenue_display,
                            revenue_change=stats.get('revenue_change', 0),
                            expiring_count=stats.get('expiring_count', 0),
                            
                            # Helper data
                            current_month=current_month,
                            available_months=available_months,
                            gym_details=gym.get_gym_details(),
                            
                            # Charts
                            revenue_trend=revenue_trend,
                            max_revenue=max_revenue,
                            
                            # Smart Alerts lists
                            unpaid_members_alert=alerts.get('unpaid', []),
                            expiring_trials_alert=alerts.get('expiring', []),
                            birthdays_today_alert=alerts.get('birthdays', []),
                            inactive_members_alert=alerts.get('inactive', [])
                            )
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        return f"<h1>Global Error Handler</h1><pre>{error_details}</pre>", 500


@app.route('/analytics')
def analytics():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    all_members = gym.get_all_members()
    total_members = len(all_members)
    
    # Calculate Member Growth (last 6 months)
    member_growth = []
    max_growth = 0
    current_date = datetime.now()
    
    for i in range(5, -1, -1):
        year = current_date.year
        month = current_date.month - i
        while month <= 0:
            month += 12
            year -= 1
        
        month_start = datetime(year, month, 1).date()
        if month == 12:
            month_end = datetime(year + 1, 1, 1).date()
        else:
            month_end = datetime(year, month + 1, 1).date()
        
        # Count members who joined in this month
        count = sum(1 for m in all_members 
                   if m.get('joined_date') and 
                   month_start <= datetime.strptime(str(m['joined_date']), '%Y-%m-%d').date() < month_end)
        
        member_growth.append({
            'month': datetime(year, month, 1).strftime('%b'),
            'count': count
        })
        max_growth = max(max_growth, count)
    
    # Calculate Retention Rate (members who renewed this month)
    current_month = datetime.now().strftime('%Y-%m')
    last_month = (datetime.now().replace(day=1) - timedelta(days=1)).strftime('%Y-%m')
    
    paid_this_month = sum(1 for m in all_members if gym.is_fee_paid(m['id'], current_month))
    paid_last_month = sum(1 for m in all_members if gym.is_fee_paid(m['id'], last_month))
    
    retention_rate = round((paid_this_month / paid_last_month * 100), 1) if paid_last_month > 0 else 100
    
    # Total Check-ins (this month)
    total_checkins = 0
    month_start = datetime.now().replace(day=1)
    
    for member in all_members:
        attendance = gym.get_attendance(member['id'])
        for record in attendance:
            try:
                checkin_date = datetime.strptime(record['timestamp'], '%Y-%m-%d %H:%M:%S')
                if checkin_date >= month_start:
                    total_checkins += 1
            except:
                pass
    
    # Attendance Heatmap (Day of week vs Hour)
    heatmap_data = {}
    max_attendance = 0
    
    for member in all_members:
        attendance = gym.get_attendance(member['id'])
        for record in attendance:
            try:
                checkin_time = datetime.strptime(record['timestamp'], '%Y-%m-%d %H:%M:%S')
                day_of_week = checkin_time.weekday()  # 0=Monday, 6=Sunday
                hour = checkin_time.hour
                
                key = (day_of_week, hour)
                heatmap_data[key] = heatmap_data.get(key, 0) + 1
                max_attendance = max(max_attendance, heatmap_data[key])
            except:
                pass
    
    # Top Performers (Most active members)
    member_checkins = []
    for member in all_members:
        attendance = gym.get_attendance(member['id'])
        checkin_count = len(attendance)
        if checkin_count > 0:
            member_checkins.append({
                'name': member['name'],
                'phone': member['phone'],
                'checkins': checkin_count
            })
    
    # Sort and get top 5
    top_performers = sorted(member_checkins, key=lambda x: x['checkins'], reverse=True)[:5]
    
    return render_template('analytics.html',
                         total_members=total_members,
                         retention_rate=retention_rate,
                         total_checkins=total_checkins,
                         member_growth=member_growth,
                         max_growth=max_growth if max_growth > 0 else 1,
                         heatmap_data=heatmap_data,
                         max_attendance=max_attendance if max_attendance > 0 else 1,
                         top_performers=top_performers)

@app.route('/advanced-analytics')
@app.route('/advanced_analytics')
def advanced_analytics():
    """Advanced Analytics optimized with Batch Data Engine"""
    gate = _require_feature('advanced_analytics', 'Advanced Analytics is available in Professional tier and above')
    if gate:
        return gate

    gym_man = get_gym()
    if not gym_man: return redirect(url_for('auth'))
    
    # Use high-performance batch engine
    data = gym_man.get_batch_analytics_data()
    if not data or not data['members']:
        flash('No data available for analytics yet.', 'info')
        return redirect(url_for('dashboard'))
        
    metrics = gym_man.calculate_business_metrics(data)
    
    # Additional data for UI charts
    growth_months = metrics['forecast_months'][:6]
    new_members_data = [] # Logic moved to metrics in future, currently 0 for churn logic
    for i in range(5, -1, -1):
        date = datetime.now() - timedelta(days=30*i)
        month_str = date.strftime('%Y-%m')
        new_members_data.append(sum(1 for m in data['members'] if m.joined_date and m.joined_date.strftime('%Y-%m') == month_str))
    
    churned_members_data = metrics['churn_trend']
    membership_types = ['Gym', 'Gym + Cardio', 'Personal Training', 'Other']
    revenue_by_type = [0] * 4
    
    # Calculate revenue by type in-memory from batch fees
    current_month = datetime.now().strftime('%Y-%m')
    member_map = {m.id: m for m in data['members']}
    for fee in data['fees']:
        if fee.month == current_month:
            m = member_map.get(fee.member_id)
            if m:
                try:
                    idx = membership_types.index(m.membership_type)
                    revenue_by_type[idx] += float(fee.amount)
                except:
                    revenue_by_type[3] += float(fee.amount)
                    
    # Intelligent Insights
    insights = []
    if metrics['revenue_growth'] > 10:
        insights.append({'type': 'success', 'icon': '📈', 'title': 'Revenue Surge', 'message': f'Revenue increased by {metrics["revenue_growth"]}%! Your growth strategy is working.'})
    elif metrics['revenue_growth'] < -10:
        insights.append({'type': 'warning', 'icon': '⚠️', 'title': 'Revenue Warning', 'message': f'Revenue dropped by {abs(metrics["revenue_growth"])}%. Check for churn or late payments.'})
        
    if metrics['at_risk_count'] > 5:
        insights.append({'type': 'danger', 'icon': '🚨', 'title': 'At-Risk Members', 'message': f'{metrics["at_risk_count"]} members are at risk of churning (no attendance in 7 days).'})

    return render_template('advanced_analytics.html',
                         total_revenue=metrics['total_revenue'],
                         revenue_growth=metrics['revenue_growth'],
                         member_growth=metrics.get('member_growth', 0),
                         total_members=len(data['members']),
                         new_members_this_month=new_members_data[-1],
                         retention_rate=metrics['retention_rate'],
                         avg_attendance=round(len(data['attendance']) / 30, 1),
                         peak_hour=metrics['peak_hour'],
                         forecast_months=metrics['forecast_months'],
                         actual_revenue=metrics['actual_revenue'],
                         forecasted_revenue=metrics['forecasted_revenue'],
                         growth_months=growth_months,
                         new_members_data=new_members_data,
                         churned_members_data=churned_members_data,
                         membership_types=membership_types,
                         revenue_by_type=revenue_by_type,
                         collection_months=metrics['collection_months'],
                         collection_rates=metrics['collection_rates'],
                         heatmap_hours=metrics['heatmap_hours'],
                         heatmap_data=metrics['heatmap_data'],
                            insights=insights,
                         vip_count=metrics['vip_count'],
                         active_count=metrics['active_count'],
                         at_risk_count=metrics['at_risk_count'],
                         churned_count=metrics['churned_count'])
    
@app.route('/bulk-operations')
def bulk_operations():
    gate = _require_feature('bulk_operations', 'Bulk Operations is available in Professional tier and above')
    if gate:
        return gate

    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    all_members = gym.get_all_members()
    current_month = datetime.now().strftime('%Y-%m')
    
    # Add is_paid status to members
    for member in all_members:
        member['is_paid'] = gym.is_fee_paid(member['id'], current_month)
    
    # Generate months (no fixed past limit)
    available_months = _build_available_months(gym=gym, as_dict=False)
    
    return render_template('bulk_operations.html',
                         members=all_members,
                         current_month=current_month,
                         available_months=available_months)

@app.route('/bulk-payment', methods=['POST'])
def bulk_payment():
    gate = _require_feature('bulk_operations', 'Bulk Operations is available in Professional tier and above')
    if gate:
        return gate

    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member_ids = request.form.get('member_ids', '').split(',')
    month = request.form.get('month')
    amount = float(request.form.get('amount', 0))
    
    success_count = 0
    for member_id in member_ids:
        if member_id.strip():
            if gym.pay_fee(member_id.strip(), month, amount, datetime.now().strftime('%Y-%m-%d')):
                success_count += 1
    
    flash(f'✅ Successfully recorded {success_count} payments!', 'success')
    return redirect(url_for('bulk_operations'))



@app.route('/add_member', methods=['GET', 'POST'])
def add_member():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    # Generate months for dropdown (no fixed past limit, keep future months)
    available_months = _build_available_months(gym=gym, future_months=24, as_dict=True)
    
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        membership_type = request.form.get('membership_type', 'Gym')
        
        # Initial Payment data
        initial_month = request.form.get('initial_month')
        try:
            initial_amount = float(request.form.get('initial_amount', 0) or 0)
        except ValueError:
            initial_amount = 0
            
        photo_path = None
        
        # Handle file upload
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                photo_path = filename
        
        # Handle camera capture (base64 data)
        elif 'camera_photo' in request.form and request.form['camera_photo']:
            import base64
            photo_data = request.form['camera_photo'].split(',')[1]
            photo_bytes = base64.b64decode(photo_data)
            filename = f"camera_{datetime.now().strftime('%Y%m%d%H%M%S')}.png"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            with open(filepath, 'wb') as f:
                f.write(photo_bytes)
            photo_path = filename
        
        try:
            membership_type = request.form.get('membership_type', 'Gym')
            joined_date = request.form.get('joined_date')
            email = request.form.get('email')
            start_trial = request.form.get('start_trial') == 'on'
            
            # If initial payment overrides trial, we can decide logic.
            # Here: If they pay, trial is False. If they don't and check trial, it's True.
            if initial_amount > 0:
                start_trial = False
                
            member_id = gym.add_member(name, phone, photo_path, membership_type, joined_date, is_trial=start_trial, email=email)
            
            # Record initial payment if amount > 0
            if initial_amount > 0 and initial_month:
                gym.pay_fee(member_id, initial_month, initial_amount)
                flash(f'Member {name} added and payment recorded for {initial_month}!', 'success')
            elif start_trial:
                flash(f'Member {name} added on 3-Day Free Trial! 🆓', 'success')
            else:
                flash(f'Member {name} added successfully! (ID: {member_id})', 'success')
                
            return redirect(url_for('dashboard'))
            
        except Exception as e:
            flash(f'Error adding member: {str(e)}', 'error')
            return redirect(url_for('add_member'))
    
    current_date = datetime.now()
    return render_template('add_member.html', 
                         available_months=available_months, 
                         current_month=current_date.strftime('%Y-%m'),
                         today=current_date.strftime('%Y-%m-%d'))

@app.route('/fees', methods=['GET', 'POST'])
def fees():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if request.method == 'POST':
        member_id = request.form.get('member_id')
        month = request.form.get('month')
        amount = float(request.form.get('amount') or 0)
        notes = request.form.get('notes', '')
        
        if gym.pay_fee(member_id, month, amount, notes):
            member = gym.get_member(member_id)
            flash(f'Fee recorded for {member["name"]} for {month}', 'success')
        else:
            flash('Member not found!', 'error')
        
        return redirect(url_for('fees'))
    
    # Get current month
    current_month = datetime.now().strftime('%Y-%m')

    fees_list = []
    total_collected = 0
    current_month_total = 0
    all_members = []
    paid_members = []
    unpaid_members = []

    try:
        # Get all payment records across all members for this gym
        if getattr(gym, 'legacy', False):
            members_by_id = {str(member.get('id')): member for member in gym.get_all_members()}
            legacy_fees = gym.data.get('fees', {}) if hasattr(gym, 'data') else {}

            for member_id, month_records in legacy_fees.items():
                member = members_by_id.get(str(member_id))
                if not member:
                    continue
                for month_value, info in month_records.items():
                    amount_value = float(info.get('amount', 0) or 0)
                    raw_date = info.get('date') or info.get('timestamp')
                    paid_date = None
                    if raw_date:
                        try:
                            paid_date = datetime.strptime(raw_date, '%Y-%m-%d %H:%M:%S')
                        except Exception:
                            try:
                                paid_date = datetime.strptime(raw_date, '%Y-%m-%d')
                            except Exception:
                                paid_date = datetime.now()
                    else:
                        paid_date = datetime.now()

                    fees_list.append({
                        'member_id': str(member_id),
                        'member': {
                            'name': member.get('name', 'Unknown'),
                            'phone': member.get('phone', 'N/A')
                        },
                        'month': month_value,
                        'amount': amount_value,
                        'paid_date': paid_date,
                        'notes': info.get('notes', '')
                    })

            fees_list.sort(key=lambda fee: fee['paid_date'], reverse=True)
        else:
            fees_query = gym.session.query(Fee).join(Member).filter(Member.gym_id == gym.gym.id).order_by(Fee.paid_date.desc())
            fees_list = fees_query.all()

        # Calculate summary
        total_collected = sum(float(f.amount or 0) for f in fees_list)
        current_month_records = [f for f in fees_list if f.month == current_month]
        current_month_total = sum(float(f.amount or 0) for f in current_month_records)

        # Get all members for the dropdown
        all_members = gym.get_all_members()

        # Get paid/unpaid members for current month
        for member in all_members:
            is_paid = gym.is_fee_paid(member['id'], current_month)
            if is_paid:
                paid_members.append(member)
            else:
                unpaid_members.append(member)
    except Exception as fees_error:
        print(f"❌ Fees route error: {fees_error}")
        traceback.print_exc()
        flash('Some payment data could not be loaded. Showing available information.', 'warning')
    
    # Generate months for dropdown (no fixed past limit, keep future months)
    available_months = _build_available_months(gym=gym, future_months=24, as_dict=True)
    
    return render_template('fees.html', 
                         members=all_members,
                         paid_members=paid_members,
                         unpaid_members=unpaid_members,
                         fees=fees_list,
                         current_month=current_month,
                         available_months=available_months,
                         total_collected=total_collected,
                         current_month_total=current_month_total,
                         gym_details=gym.get_gym_details())

@app.route('/download_excel')
def download_excel():
    """Quick export - dashboard members list"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    current_month = datetime.now().strftime('%Y-%m')
    all_members = gym.get_all_members()
    
    # Prepare comprehensive data for Excel
    data = []
    for member in all_members:
        # Check payment status
        is_paid = gym.is_fee_paid(member['id'], current_month)
        
        data.append({
            'ID': member['id'],
            'Name': member['name'],
            'Phone': member['phone'],
            'Email': member.get('email', 'N/A'),
            'Membership Type': member.get('membership_type', 'Gym'),
            'Join Date': member.get('joined_date', 'N/A'),
            'Status': 'PAID' if is_paid else 'UNPAID',
            'Active': 'Yes' if member.get('active', True) else 'No'
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file with formatting
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Members')
        
        # Get workbook and worksheet
        worksheet = writer.sheets['Members']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    output.seek(0)
    
    filename = f'gym_members_{current_month}.xlsx'
    return send_file(output, download_name=filename, as_attachment=True)

# ==================== ADVANCED EXPORT CENTER ====================

@app.route('/export_center')
def export_center():
    """Export Center - Dashboard for all export options"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    return render_template('export_center.html')

@app.route('/export/members_complete')
def export_members_complete():
    """Export complete member database"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    from export_manager import ExportManager
    export_mgr = ExportManager(gym)
    
    output = export_mgr.export_members_complete()
    filename = f'members_complete_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/revenue_report')
def export_revenue_report():
    """Export revenue report"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    from export_manager import ExportManager
    export_mgr = ExportManager(gym)
    
    # Get date range from query params
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    output = export_mgr.export_revenue_report(start_date, end_date)
    filename = f'revenue_report_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/export/unpaid_members')
def export_unpaid_members():
    """Export unpaid members list"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    from export_manager import ExportManager
    export_mgr = ExportManager(gym)
    
    output = export_mgr.export_unpaid_members()
    filename = f'unpaid_members_{datetime.now().strftime("%Y%m%d")}.xlsx'
    
    return send_file(output, download_name=filename, as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/card/<member_id>')
def generate_card(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member = gym.get_member(member_id)
    if not member:
        flash('Member not found!', 'error')
        return redirect(url_for('dashboard'))
    
    # Create PDF
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    # Card background
    c.setFillColorRGB(0.1, 0.1, 0.2)
    c.rect(50, height - 350, 300, 200, fill=True)
    
    # Title
    c.setFillColorRGB(1, 1, 1)
    c.setFont("Helvetica-Bold", 20)
    c.drawString(70, height - 180, "GYM MEMBER CARD")
    
    # Member photo
    if member.get('photo'):
        photo_path = os.path.join(app.config['UPLOAD_FOLDER'], member['photo'])
        if os.path.exists(photo_path):
            try:
                img = ImageReader(photo_path)
                c.drawImage(img, 70, height - 330, width=80, height=100, preserveAspectRatio=True)
            except:
                pass
    
    # QR Code
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(member_id)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    
    # Save QR to buffer to draw
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer)
    qr_buffer.seek(0)
    
    # Draw QR Code on card
    c.drawImage(ImageReader(qr_buffer), 270, height - 330, width=70, height=70)
    
    # Member details
    c.setFont("Helvetica", 12)
    c.drawString(170, height - 230, f"ID: {member['id']}")
    c.drawString(170, height - 250, f"Name: {member['name']}")
    c.drawString(170, height - 270, f"Phone: {member['phone']}")
    c.drawString(170, height - 290, f"Joined: {member['joined_date']}")
    
    c.save()
    buffer.seek(0)
    
    return send_file(buffer, download_name=f'card_{member_id}.pdf', as_attachment=True, mimetype='application/pdf')

@app.route('/scanner')
def scanner():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    return render_template('scanner.html')

@app.route('/scan_check/<member_id>')
def scan_check(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    # Determine status
    current_month = datetime.now().strftime('%Y-%m')
    is_paid = gym.is_fee_paid(member_id, current_month)
    member = gym.get_member(member_id)
    
    if not member:
        flash('Invalid Member ID!', 'error')
        return redirect(url_for('scanner'))
        
    status = 'GRANTED' if is_paid else 'DENIED'
    status = ''
    if is_paid:
        status = 'ACCESS GRANTED'
        # Log attendance automatically
        gym.log_attendance(member_id)
        
        # Check for milestone alerts (Phase 2)
        try:
            attn_history = gym.get_attendance(member_id)
            total_visits = len(attn_history)
            if total_visits in [50, 100, 250, 500, 1000]:
                session_db = get_session()
                auto_man = AutomationManager(session_db, email_sender)
                auto_man.send_milestone_alert(member_id, total_visits)
                session_db.close()
        except Exception as e:
            print(f"⚠️ Milestone alert error: {str(e)}")
    # Special check for trial
    elif not is_paid and member.get('is_trial'):
        today = datetime.now().strftime('%Y-%m-%d')
        if member.get('trial_end_date') >= today:
             status = 'TRIAL'
        else:
            status = 'ACCESS DENIED - TRIAL EXPIRED'
    else:
        status = 'ACCESS DENIED - FEE PENDING'
    
    # Get attendance history
    attendance_history = gym.get_attendance(member_id)
    
    # Get payment details
    payment_history = gym.get_member_fees(member_id)
    last_payment = payment_history[0] if payment_history else None
             
    return render_template('scan_result.html', 
                         member=member, 
                         status=status, 
                         month=current_month,
                         attendance_history=attendance_history,
                         last_payment=last_payment,
                         is_paid=is_paid,
                         gym_details=gym.get_gym_details())

@app.route('/member/<member_id>', methods=['GET', 'POST'])
def member_details(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member = gym.get_member(member_id)
    if not member:
        flash('Member not found!', 'error')
        return redirect(url_for('dashboard'))
        
    # Safe data fetching to prevent 500 errors if DB schema is mismatched
    attendance_history = []
    fees_history = []
    try:
        attendance_history = gym.get_attendance(member_id)
        fees_history = gym.get_member_fees(member_id)
    except Exception as e:
        print(f"⚠️ DB Error fetching details: {str(e)}")
        flash(f'Database Schema Error: {str(e)}', 'error')
        flash('⚠️ Please run the Database Fix Tool to repair this!', 'warning')
        # We don't return here, we let the page load with empty history so user can see the "Fix DB" button
    
    if request.method == 'POST':
        month = request.form.get('month')
        amount = float(request.form.get('amount') or 0)
        payment_date = request.form.get('payment_date')
        notes = request.form.get('notes')
        
        if gym.pay_fee(member_id, month, amount, payment_date, notes):
            flash(f'Payment recorded successfully for {month}!', 'success')
        else:
            flash('Payment failed!', 'error')
        return redirect(url_for('member_details', member_id=member_id))
    
    
    available_months = _build_available_months(gym=gym, as_dict=True)
    
    return render_template('member_details.html', 
                         member=member, 
                         gym_details=gym.get_gym_details(), 
                         history=fees_history,
                         attendance_history=attendance_history,
                         current_month=datetime.now().strftime('%Y-%m'),
                         today=datetime.now().strftime('%Y-%m-%d'),
                         available_months=available_months,
                         notes=gym.get_member_notes(member_id),
                         timeline=gym.get_member_timeline(member_id),
                         measurements=gym.get_body_measurements(member_id))

@app.route('/member/<member_id>/add_note', methods=['POST'])
def add_member_note(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    note_text = request.form.get('note')
    if note_text and note_text.strip():
        if gym.add_member_note(member_id, note_text.strip()):
            flash('📝 Note added successfully!', 'success')
        else:
            flash('Failed to add note', 'error')
    
    return redirect(url_for('member_details', member_id=member_id))

@app.route('/member/<member_id>/delete_note/<note_id>')
def delete_member_note(member_id, note_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if gym.delete_member_note(note_id):
        flash('🗑️ Note deleted!', 'success')
    else:
        flash('Failed to delete note', 'error')
    
    return redirect(url_for('member_details', member_id=member_id))

@app.route('/member/<member_id>/add-measurement', methods=['POST'])
def add_measurement(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    weight = request.form.get('weight')
    body_fat = request.form.get('body_fat')
    chest = request.form.get('chest')
    waist = request.form.get('waist')
    arms = request.form.get('arms')
    notes = request.form.get('notes')
    
    if weight:
        if gym.add_body_measurement(member_id, weight, body_fat, chest, waist, arms, notes):
            flash('📊 Measurement recorded successfully!', 'success')
        else:
            flash('Failed to add measurement', 'error')
    
    return redirect(url_for('member_details', member_id=member_id))

@app.route('/member/<member_id>/delete-measurement/<measurement_id>')
def delete_measurement(member_id, measurement_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if gym.delete_body_measurement(measurement_id):
        flash('🗑️ Measurement deleted!', 'success')
    else:
        flash('Failed to delete measurement', 'error')
    
    return redirect(url_for('member_details', member_id=member_id))



@app.route('/member/<member_id>/delete_fee/<month>', methods=['POST'])
def delete_fee_record(member_id, month):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if gym.delete_fee(member_id, month):
        flash(f'Payment for {month} deleted!', 'success')
    else:
        flash('Delete failed!', 'error')
        
    return redirect(url_for('member_details', member_id=member_id))

@app.route('/member/<member_id>/edit_fee/<month>', methods=['GET', 'POST'])
def edit_fee_record(member_id, month):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member = gym.get_member(member_id)
    if not member or not gym.is_fee_paid(member_id, month):
        flash('Fee record not found!', 'error')
        return redirect(url_for('member_details', member_id=member_id))
        
    if request.method == 'POST':
        try:
            amount = float(request.form.get('amount') or 0)
            date = request.form.get('date') # Expecting YYYY-MM-DD HH:MM:SS or just date
            
            if gym.update_fee(member_id, month, amount, date):
                flash(f'Payment for {month} updated!', 'success')
                return redirect(url_for('member_details', member_id=member_id))
            else:
                flash('Update failed!', 'error')
        except ValueError:
            flash('Invalid amount!', 'error')
            
    # Get current fee data
    fee_info = None
    if gym.legacy and hasattr(gym, 'data'):
        if member_id in gym.data.get('fees', {}) and month in gym.data['fees'][member_id]:
            fee_info = gym.data['fees'][member_id][month]
    else:
        # Database mode - get from Fee table
        try:
            # Use the gym's existing session if available, or create new one properly
            if hasattr(gym, 'session'):
                session = gym.session
            else:
                from models import get_session
                session = get_session()
                
            if session:
                from models import Fee
                fee = session.query(Fee).filter_by(member_id=int(member_id), month=month).first()
                if fee:
                    fee_info = {
                        'amount': fee.amount,
                        'date': fee.paid_date.strftime('%Y-%m-%d %H:%M:%S') if fee.paid_date else '',
                        'timestamp': fee.paid_date.strftime('%Y-%m-%d %H:%M:%S') if fee.paid_date else ''
                    }
        except Exception as e:
            flash(f"Error retrieving fee: {str(e)}", "error")
    
    if not fee_info:
        flash('Fee record not found', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('edit_fee.html', member=member, month=month, fee=fee_info)

@app.route('/member/<member_id>/edit', methods=['GET', 'POST'])
def edit_member(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member = gym.get_member(member_id)
    if not member:
        flash('Member not found!', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name')
        phone = request.form.get('phone')
        email = request.form.get('email')
        membership_type = request.form.get('membership_type')
        joined_date = request.form.get('joined_date')
        
        if gym.update_member(member_id, name, phone, membership_type, joined_date, email):
            flash('Member updated successfully!', 'success')
            return redirect(url_for('member_details', member_id=member_id))
        else:
            flash('Update failed!', 'error')
    
    return render_template('edit_member.html', member=member)

@app.route('/member/<member_id>/delete', methods=['POST'])
def delete_member(member_id):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if gym.delete_member(member_id):
        # Audit Log
        user_id = auth_manager.get_user_id(session.get('username'))
        if user_id:
            security_manager.log_action(user_id, 'MEMBER_DELETE', {'member_id': member_id}, request.remote_addr, request.user_agent.string)
        
        flash('Member deleted successfully!', 'success')
    else:
        flash('Delete failed!', 'error')
        
    return redirect(url_for('dashboard'))

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))

    current_user = None
    if not auth_manager.legacy and session.get('username'):
        current_user = auth_manager.session.query(User).filter_by(email=session['username']).first()
    
    if request.method == 'POST':
        action = request.form.get('action', '').strip()

        if action == 'share_staff_access':
            if auth_manager.legacy:
                flash('Staff sharing requires database mode.', 'error')
                return redirect(url_for('settings'))

            if not current_user:
                flash('Session expired. Please login again.', 'error')
                return redirect(url_for('auth'))

            staff_email = (request.form.get('staff_email') or '').strip().lower()
            if not staff_email or '@' not in staff_email:
                flash('Please enter a valid staff email address.', 'error')
                return redirect(url_for('settings'))

            if staff_email == current_user.email.lower():
                flash('You cannot share staff access with your own email.', 'error')
                return redirect(url_for('settings'))

            owner_gym = auth_manager.session.query(Gym).filter_by(user_id=current_user.id).first()
            if not owner_gym:
                flash('No gym found for your account. Please update gym settings first.', 'error')
                return redirect(url_for('settings'))

            staff_user = auth_manager.session.query(User).filter_by(email=staff_email).first()
            temp_password = None
            if not staff_user:
                temp_password = secrets.token_urlsafe(8)
                staff_user = User(
                    email=staff_email,
                    password_hash=auth_manager.hash_password(temp_password),
                    role='staff',
                    market=current_user.market,
                    subscription_expiry=current_user.subscription_expiry,
                    subscription_tier=current_user.subscription_tier,
                    billing_cycle=current_user.billing_cycle,
                    subscription_status=current_user.subscription_status
                )
                auth_manager.session.add(staff_user)
                auth_manager.session.flush()
            else:
                staff_user.role = 'staff'

            existing_access = auth_manager.session.query(StaffAccess).filter_by(
                staff_user_id=staff_user.id,
                owner_user_id=current_user.id,
                gym_id=owner_gym.id
            ).first()

            tier_config = TierManager.get_tier_config(current_user.subscription_tier or 'starter')
            staff_limit = int((tier_config.get('limits') or {}).get('staff', -1))
            active_staff_count = auth_manager.session.query(StaffAccess).filter_by(
                owner_user_id=current_user.id,
                gym_id=owner_gym.id,
                is_active=True
            ).count()

            projected_count = active_staff_count
            if not existing_access or not existing_access.is_active:
                projected_count += 1

            if staff_limit != -1 and projected_count > staff_limit:
                flash(
                    f"Staff seat limit reached for your {tier_config.get('name', 'Starter')} plan ({active_staff_count}/{staff_limit}). Upgrade your plan to add more staff.",
                    'warning'
                )
                return redirect(url_for('settings'))

            if existing_access:
                existing_access.is_active = True
            else:
                auth_manager.session.add(StaffAccess(
                    staff_user_id=staff_user.id,
                    owner_user_id=current_user.id,
                    gym_id=owner_gym.id,
                    is_active=True
                ))

            auth_manager.session.commit()

            login_url = request.host_url.rstrip('/') + url_for('auth')
            subject = f"Staff Access Granted - {owner_gym.name}"
            if temp_password:
                body = f"""
                <h3>Staff Access Granted</h3>
                <p>You have been granted staff access to <strong>{owner_gym.name}</strong>.</p>
                <p><strong>Login Email:</strong> {staff_email}<br>
                <strong>Temporary Password:</strong> {temp_password}</p>
                <p>Please login at: <a href=\"{login_url}\">{login_url}</a> and change your password immediately.</p>
                """
            else:
                body = f"""
                <h3>Staff Access Updated</h3>
                <p>Your account now has staff access to <strong>{owner_gym.name}</strong>.</p>
                <p>Login at: <a href=\"{login_url}\">{login_url}</a></p>
                """

            try:
                email_sent = email_sender.send_email(staff_email, subject, body)
                if email_sent:
                    flash(f'Staff access shared with {staff_email} and email notification sent.', 'success')
                else:
                    flash(
                        f'Staff access shared with {staff_email}, but email was not sent. Configure SMTP in environment settings to enable notifications.',
                        'warning'
                    )
            except Exception:
                flash(
                    f'Staff access shared with {staff_email}, but email failed. Configure SMTP credentials to enable notifications.',
                    'warning'
                )

            try:
                security_manager.log_action(
                    current_user.id,
                    'STAFF_ACCESS_SHARED',
                    {
                        'staff_email': staff_email,
                        'gym_name': owner_gym.name,
                        'email_sent': bool(email_sent) if 'email_sent' in locals() else False
                    },
                    request.remote_addr,
                    request.user_agent.string if request.user_agent else None
                )
            except Exception:
                pass

            return redirect(url_for('settings'))

        if action == 'revoke_staff_access':
            if auth_manager.legacy:
                flash('Staff sharing requires database mode.', 'error')
                return redirect(url_for('settings'))

            if not current_user:
                flash('Session expired. Please login again.', 'error')
                return redirect(url_for('auth'))

            access_id = request.form.get('access_id')
            if not access_id:
                flash('Invalid revoke request.', 'error')
                return redirect(url_for('settings'))

            try:
                access_id_int = int(access_id)
            except Exception:
                flash('Invalid revoke request.', 'error')
                return redirect(url_for('settings'))

            access = auth_manager.session.query(StaffAccess).filter_by(
                id=access_id_int,
                owner_user_id=current_user.id,
                is_active=True
            ).first()

            if not access:
                flash('Staff access record not found.', 'error')
                return redirect(url_for('settings'))

            staff_user = auth_manager.session.query(User).filter_by(id=access.staff_user_id).first()
            access_gym = auth_manager.session.query(Gym).filter_by(id=access.gym_id).first()
            gym_name = access_gym.name if access_gym else 'your gym'

            access.is_active = False

            remaining_active = auth_manager.session.query(StaffAccess).filter_by(
                staff_user_id=access.staff_user_id,
                is_active=True
            ).count()

            auth_manager.session.commit()

            if staff_user and staff_user.email:
                try:
                    login_url = request.host_url.rstrip('/') + url_for('auth')
                    email_sender.send_email(
                        staff_user.email,
                        f"Staff Access Revoked - {gym_name}",
                        f"""
                        <h3>Staff Access Revoked</h3>
                        <p>Your staff access to <strong>{gym_name}</strong> has been revoked.</p>
                        <p>If this is unexpected, please contact your gym owner.</p>
                        <p>Login page: <a href=\"{login_url}\">{login_url}</a></p>
                        """
                    )
                except Exception:
                    pass

            if remaining_active > 0:
                flash('Staff access revoked successfully. User still has access to other shared gyms.', 'success')
            else:
                flash('Staff access revoked successfully.', 'success')

            try:
                security_manager.log_action(
                    current_user.id,
                    'STAFF_ACCESS_REVOKED',
                    {
                        'staff_email': staff_user.email if staff_user else None,
                        'gym_name': gym_name,
                        'has_other_active_access': remaining_active > 0
                    },
                    request.remote_addr,
                    request.user_agent.string if request.user_agent else None
                )
            except Exception:
                pass

            return redirect(url_for('settings'))

        if action == 'resend_staff_invite':
            if auth_manager.legacy:
                flash('Staff sharing requires database mode.', 'error')
                return redirect(url_for('settings'))

            if not current_user:
                flash('Session expired. Please login again.', 'error')
                return redirect(url_for('auth'))

            access_id = request.form.get('access_id')
            if not access_id:
                flash('Invalid resend request.', 'error')
                return redirect(url_for('settings'))

            try:
                access_id_int = int(access_id)
            except Exception:
                flash('Invalid resend request.', 'error')
                return redirect(url_for('settings'))

            access = auth_manager.session.query(StaffAccess).filter_by(
                id=access_id_int,
                owner_user_id=current_user.id,
                is_active=True
            ).first()

            if not access:
                flash('Staff access record not found.', 'error')
                return redirect(url_for('settings'))

            staff_user = auth_manager.session.query(User).filter_by(id=access.staff_user_id).first()
            owner_gym = auth_manager.session.query(Gym).filter_by(id=access.gym_id).first()

            if not staff_user or not staff_user.email:
                flash('Staff user email not found.', 'error')
                return redirect(url_for('settings'))

            login_url = request.host_url.rstrip('/') + url_for('auth')
            reset_url = request.host_url.rstrip('/') + url_for('forgot_password')
            gym_name = owner_gym.name if owner_gym else 'your gym'

            try:
                email_sent = email_sender.send_email(
                    staff_user.email,
                    f"Staff Access Reminder - {gym_name}",
                    f"""
                    <h3>Staff Access Reminder</h3>
                    <p>You have active staff access to <strong>{gym_name}</strong>.</p>
                    <p>Login page: <a href=\"{login_url}\">{login_url}</a></p>
                    <p>If you forgot your password, reset it here: <a href=\"{reset_url}\">{reset_url}</a></p>
                    """
                )
                if email_sent:
                    flash(f'Reminder sent to {staff_user.email}.', 'success')
                else:
                    flash(
                        f'Could not send reminder to {staff_user.email}. Configure SMTP in environment settings.',
                        'warning'
                    )
            except Exception:
                flash(
                    f'Failed to send reminder to {staff_user.email}. Configure SMTP credentials to enable notifications.',
                    'warning'
                )

            try:
                security_manager.log_action(
                    current_user.id,
                    'STAFF_ACCESS_RESEND',
                    {
                        'staff_email': staff_user.email,
                        'gym_name': gym_name,
                        'email_sent': bool(email_sent) if 'email_sent' in locals() else False
                    },
                    request.remote_addr,
                    request.user_agent.string if request.user_agent else None
                )
            except Exception:
                pass

            return redirect(url_for('settings'))

        name = request.form.get('gym_name')
        currency = request.form.get('currency', '$')
        if currency == 'AUTO':
            currency = ''
        logo_path = None
        
        if 'gym_logo' in request.files:
            file = request.files['gym_logo']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"logo_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)
                logo_path = filename
        
        if gym.update_gym_details(name, logo_path, currency):
            flash('Gym settings updated successfully!', 'success')
        else:
            flash('Failed to update settings!', 'error')
        return redirect(url_for('settings'))

    payments = []
    subscription = {
        'market': 'US',
        'status': 'Active',
        'expiry': 'Lifetime',
        'days_left': 9999,
        'is_trial': False,
        'plan_name': 'Starter'
    }
    
    if 'logged_in' in session:
        username = session['username']
        if auth_manager.legacy:
            legacy_user = auth_manager.users.get(username, {})
            payments = legacy_user.get('payments', [])
            subscription = {
                'market': 'US',
                'status': 'Active',
                'expiry': 'Lifetime',
                'days_left': 9999,
                'is_trial': False,
                'plan_name': 'Starter'
            }
        else:
            user = auth_manager.session.query(User).filter_by(email=username).first()

            if user:
                is_active = auth_manager.is_subscription_active(username)
                days_left = 0
                if user.subscription_expiry:
                    days_left = (user.subscription_expiry - datetime.utcnow()).days

                subscription = {
                    'market': user.market or 'US',
                    'status': 'Active' if is_active else 'Expired',
                    'expiry': user.subscription_expiry.strftime('%Y-%m-%d') if user.subscription_expiry else 'Lifetime',
                    'days_left': days_left,
                    'is_trial': False, # Simplify for now
                    'plan_name': TierManager.get_tier_config(user.subscription_tier or 'starter').get('name', 'Starter')
                }
        
    shared_staff = []
    staff_activity_logs = []
    staff_limit_info = {
        'used': 0,
        'limit': 0,
        'is_unlimited': False,
        'plan_name': subscription.get('plan_name', 'Starter')
    }
    if not auth_manager.legacy and current_user:
        owner_gym = auth_manager.session.query(Gym).filter_by(user_id=current_user.id).first()
        if owner_gym:
            accesses = auth_manager.session.query(StaffAccess, User).join(
                User, User.id == StaffAccess.staff_user_id
            ).filter(
                StaffAccess.owner_user_id == current_user.id,
                StaffAccess.gym_id == owner_gym.id,
                StaffAccess.is_active == True
            ).all()

            shared_staff = [{
                'access_id': access.id,
                'email': user.email,
                'role': user.role,
                'created_at': access.created_at.strftime('%Y-%m-%d') if access.created_at else ''
            } for access, user in accesses]

            staff_limit_value = TierManager.get_limit(current_user, 'staff')
            staff_limit_info = {
                'used': len(shared_staff),
                'limit': staff_limit_value,
                'is_unlimited': staff_limit_value == -1,
                'plan_name': TierManager.get_tier_config(current_user.subscription_tier or 'starter').get('name', 'Starter')
            }

            try:
                audit_logs = security_manager.get_audit_logs(user_id=current_user.id, limit=50)
                staff_activity_logs = [
                    log for log in audit_logs
                    if log.get('action') in {'STAFF_ACCESS_SHARED', 'STAFF_ACCESS_REVOKED', 'STAFF_ACCESS_RESEND'}
                ][:10]
            except Exception:
                staff_activity_logs = []

    return render_template(
        'settings.html',
        details=gym.get_gym_details(),
        payments=payments,
        subscription=subscription,
        shared_staff=shared_staff,
        staff_activity_logs=staff_activity_logs,
        staff_limit_info=staff_limit_info
    )

@app.route('/restore_backup', methods=['POST'])
def restore_backup():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if 'backup_file' not in request.files:
        flash('No file selected!', 'error')
        return redirect(url_for('settings'))
        
    file = request.files['backup_file']
    if file.filename == '':
        flash('No file selected!', 'error')
        return redirect(url_for('settings'))
        
    ext = os.path.splitext(file.filename.lower())[1]
    if file and ext in ['.json', '.xlsx']:
        try:
            # Read backup payload
            if ext == '.json':
                data = json.load(file)
            else:
                data = _parse_excel_backup_to_legacy(file)
            
            # Simple validation check (must have 'members' key)
            if 'members' not in data:
                flash('Invalid backup file! Missing member data.', 'error')
                return redirect(url_for('settings'))
            
            # Save to current gym's data file (legacy mode only)
            if gym.legacy and hasattr(gym, 'data_file'):
                with open(gym.data_file, 'w') as f:
                    json.dump(data, f, indent=2)
                flash('✅ Data restored successfully! Please log in again.', 'success')
                return redirect(url_for('logout'))
            else:
                # Database mode: Import JSON data
                success, message = gym.import_json_data(data)
                if success:
                    flash(f'✅ {message}', 'success')
                else:
                    flash(f'⚠️ Import Error: {message}', 'error')
                return redirect(url_for('settings'))
            
        except Exception as e:
            flash(f'Error restoring data: {str(e)}', 'error')
    else:
        flash('Invalid file type! Please upload a JSON or Excel (.xlsx) backup file.', 'error')
        
    return redirect(url_for('settings'))


@app.route('/backup/download/json')
def download_backup_json():
    """Download full backup in legacy-compatible JSON format."""
    gym = get_gym()
    if not gym:
        return redirect(url_for('auth'))

    payload = _build_legacy_backup_payload(gym)
    output = BytesIO()
    output.write(json.dumps(payload, indent=2, default=str).encode('utf-8'))
    output.seek(0)

    filename = f"gym_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/json')


@app.route('/backup/download/excel')
def download_backup_excel():
    """Download full backup in Excel, aligned with legacy backup structure."""
    gym = get_gym()
    if not gym:
        return redirect(url_for('auth'))

    payload = _build_legacy_backup_payload(gym)

    from openpyxl import Workbook

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = 'backup_info'
    ws_info.append(['key', 'value'])
    for key, value in (payload.get('backup_meta') or {}).items():
        ws_info.append([key, str(value)])

    gym_details = payload.get('gym_details') or {}
    if gym_details:
        ws_info.append(['gym_name', str(gym_details.get('name', 'Gym Manager'))])
        ws_info.append(['currency', str(gym_details.get('currency', '$'))])

    ws_members = wb.create_sheet('members')
    ws_members.append(['id', 'name', 'phone', 'email', 'photo', 'membership_type', 'joined_date', 'is_trial', 'trial_end_date', 'is_active'])
    for member_id, member in (payload.get('members') or {}).items():
        ws_members.append([
            member_id,
            member.get('name'),
            member.get('phone'),
            member.get('email'),
            member.get('photo'),
            member.get('membership_type'),
            member.get('joined_date'),
            member.get('is_trial'),
            member.get('trial_end_date'),
            member.get('is_active')
        ])

    ws_fees = wb.create_sheet('fees')
    ws_fees.append(['member_id', 'month', 'amount', 'date'])
    for member_id, fee_map in (payload.get('fees') or {}).items():
        for month, info in (fee_map or {}).items():
            ws_fees.append([member_id, month, info.get('amount'), info.get('date')])

    ws_attendance = wb.create_sheet('attendance')
    ws_attendance.append(['member_id', 'timestamp', 'emotion', 'confidence'])
    for member_id, logs in (payload.get('attendance') or {}).items():
        for row in (logs or []):
            ws_attendance.append([member_id, row.get('timestamp'), row.get('emotion'), row.get('confidence')])

    ws_expenses = wb.create_sheet('expenses')
    ws_expenses.append(['date', 'category', 'amount', 'description'])
    for expense in (payload.get('expenses') or []):
        ws_expenses.append([
            expense.get('date'),
            expense.get('category'),
            expense.get('amount'),
            expense.get('description')
        ])

    if not gym.legacy and gym.gym:
        ws_admin = wb.create_sheet('admin_user')
        ws_admin.append(['id', 'email', 'role', 'market', 'subscription_tier', 'subscription_status', 'subscription_expiry'])
        owner = gym.session.query(User).filter_by(id=gym.gym.user_id).first()
        if owner:
            ws_admin.append([
                owner.id,
                owner.email,
                owner.role,
                owner.market,
                owner.subscription_tier,
                owner.subscription_status,
                owner.subscription_expiry.strftime('%Y-%m-%d %H:%M:%S') if owner.subscription_expiry else ''
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gym_backup_old_structure_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/backup/download/template')
def download_backup_template_excel():
    """Download blank old-structure Excel backup template."""
    gym = get_gym()
    if not gym:
        return redirect(url_for('auth'))

    from openpyxl import Workbook

    wb = Workbook()
    ws_info = wb.active
    ws_info.title = 'backup_info'
    ws_info.append(['key', 'value'])
    ws_info.append(['gym_name', gym.get_gym_details().get('name', 'Gym Manager')])
    ws_info.append(['currency', gym.get_gym_details().get('currency', '$')])
    ws_info.append(['generated_at', datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')])

    ws_members = wb.create_sheet('members')
    ws_members.append(['id', 'name', 'phone', 'email', 'photo', 'membership_type', 'joined_date', 'is_trial', 'trial_end_date', 'is_active'])
    ws_members.append(['1001', 'Sample Member', '+1234567890', 'member@example.com', '', 'Gym', '2026-03-01', False, '', True])

    ws_fees = wb.create_sheet('fees')
    ws_fees.append(['member_id', 'month', 'amount', 'date'])
    ws_fees.append(['1001', '2026-03', 1200, '2026-03-02 10:00:00'])

    ws_attendance = wb.create_sheet('attendance')
    ws_attendance.append(['member_id', 'timestamp', 'emotion', 'confidence'])
    ws_attendance.append(['1001', '2026-03-02 07:30:00', 'happy', 0.9])

    ws_expenses = wb.create_sheet('expenses')
    ws_expenses.append(['date', 'category', 'amount', 'description'])
    ws_expenses.append(['2026-03-02', 'Utilities', 350, 'Electricity bill'])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"gym_backup_template_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/merge_duplicates', methods=['POST'])
def merge_duplicates():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    try:
        merged_count = gym.merge_members()
        if merged_count > 0:
            flash(f'✅ Successfully merged {merged_count} duplicate members!', 'success')
        else:
            flash('No duplicates found.', 'info')
            
    except Exception as e:
        flash(f'Error merging duplicates: {str(e)}', 'error')
        
    return redirect(url_for('settings'))

@app.route('/receipt/<member_id>/<month>')
def generate_receipt(member_id, month):
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member = gym.get_member(member_id)
    if not member or not gym.is_fee_paid(member_id, month):
        flash('Fee record not found!', 'error')
        return redirect(url_for('member_details', member_id=member_id))
    
    # Get fee data
    fee_info = None
    if gym.legacy and hasattr(gym, 'data'):
        if member_id in gym.data.get('fees', {}) and month in gym.data['fees'][member_id]:
            fee_info = gym.data['fees'][member_id][month]
    else:
        # Database mode
        try:
            from models import get_session, Fee
            session = get_session()
            if session:
                fee = session.query(Fee).filter_by(member_id=member_id, month=month).first()
                if fee:
                    fee_info = {
                        'amount': fee.amount,
                        'date': fee.paid_date.strftime('%Y-%m-%d') if fee.paid_date else '',
                        'timestamp': fee.paid_date.strftime('%Y-%m-%d %H:%M:%S') if fee.paid_date else ''
                    }
                session.close()
        except:
            pass
    
    if not fee_info:
        flash('Fee record not found', 'error')
        return redirect(url_for('dashboard'))

    # Create PDF
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    
    gym_details = gym.get_gym_details()
    
    # Header
    c.setFont("Helvetica-Bold", 24)
    c.drawString(50, height - 50, gym_details['name'])
    
    if gym_details.get('logo'):
        logo_path = os.path.join(app.config['UPLOAD_FOLDER'], gym_details['logo'])
        if os.path.exists(logo_path):
            try:
                img = ImageReader(logo_path)
                c.drawImage(img, width - 100, height - 80, width=50, height=50, preserveAspectRatio=True)
            except:
                pass

    c.setFont("Helvetica-Bold", 18)
    c.drawString(50, height - 100, "PAYMENT RECEIPT")
    
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 130, f"Date: {datetime.now().strftime('%Y-%m-%d')}")
    c.drawString(50, height - 150, f"Receipt #: {member_id}-{month.replace('-', '')}")
    
    # Details
    y = height - 200
    c.drawString(50, y, f"Member: {member['name']} (ID: {member_id})")
    c.drawString(50, y - 20, f"Month Paid: {month}")
    c.drawString(50, y - 40, f"Amount Paid: ${fee_info['amount']}")
    c.drawString(50, y - 60, f"Payment Date: {fee_info['paid_date']}")
    
    # Footer
    c.setFont("Helvetica-Oblique", 10)
    c.drawString(50, y - 120, "Thank you for your business!")
    
    c.save()
    buffer.seek(0)
    
    return send_file(buffer, download_name=f'receipt_{member_id}_{month}.pdf', as_attachment=True, mimetype='application/pdf')

@app.route('/bulk_import', methods=['GET', 'POST'])
def bulk_import():
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    if request.method == 'POST':
        action = request.form.get('action', 'upload')
        
        # STEP 1: Upload & Preview
        if action == 'upload':
            if 'import_file' not in request.files:
                flash('No file selected!', 'error')
                return redirect(url_for('bulk_import'))
            
            file = request.files['import_file']
            if file.filename == '':
                flash('No file selected!', 'error')
                return redirect(url_for('bulk_import'))
            
            # Validate file extension
            if not (file.filename.lower().endswith('.xlsx') or file.filename.lower().endswith('.csv')):
                flash('Invalid file format! Use .xlsx or .csv files only.', 'error')
                return redirect(url_for('bulk_import'))
            
            try:
                # Save file temporarily
                upload_folder = app.config.get('UPLOAD_FOLDER', '/tmp')
                if not os.access(upload_folder, os.W_OK):
                    upload_folder = '/tmp'
                
                filename = secure_filename(f"import_{datetime.now().strftime('%Y%m%d%H%M%S')}_{file.filename}")
                filepath = os.path.join(upload_folder, filename)
                file.save(filepath)
                
                # Validate and preview import data
                from import_validator import ImportValidator
                import openpyxl
                import csv
                file_ext = os.path.splitext(filepath.lower())[1]
                
                rows_data = []

                if file_ext == '.csv':
                    try:
                        with open(filepath, 'r', encoding='utf-8-sig', newline='') as f:
                            reader = csv.DictReader(f)
                            rows_data = list(reader)
                    except UnicodeDecodeError:
                        with open(filepath, 'r', encoding='latin-1', newline='') as f:
                            reader = csv.DictReader(f)
                            rows_data = list(reader)
                else:
                    # Read Excel data
                    wb = openpyxl.load_workbook(filepath)
                    ws = wb.active
                    
                    # Convert to list of dicts
                    headers = [cell.value for cell in ws[1]]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                        rows_data.append(row_dict)
                    wb.close()
                
                # Validate
                validator = ImportValidator(gym)
                validation_results = validator.validate_import_data(rows_data)
                
                # Store in session for confirmation
                session['import_file_path'] = filepath
                session['import_validation'] = {
                    'summary': validation_results,
                    'filename': file.filename
                }
                
                return redirect(url_for('import_preview'))
                
            except Exception as e:
                flash(f'❌ Error processing file: {str(e)}', 'error')
                return redirect(url_for('bulk_import'))
        
        # STEP 2: Confirm Import (from preview)
        elif action == 'confirm':
            filepath = session.get('import_file_path')
            if not filepath or not os.path.exists(filepath):
                flash('Session expired. Please upload file again.', 'error')
                return redirect(url_for('bulk_import'))
            
            try:
                # Get duplicate strategy
                duplicate_strategy = request.form.get('duplicate_strategy', 'skip')
                
                # Process import
                success_count, error_count, errors = gym.bulk_import_members(
                    filepath, 
                    duplicate_strategy=duplicate_strategy
                )
                
                # Clean up
                try:
                    os.remove(filepath)
                    session.pop('import_file_path', None)
                    session.pop('import_validation', None)
                except:
                    pass
                
                # Show results
                if success_count > 0:
                    # Audit Log
                    user_id = auth_manager.get_user_id(session.get('username'))
                    if user_id:
                        security_manager.log_action(user_id, 'BULK_IMPORT', 
                                                   {'success_count': success_count, 'duplicate_strategy': duplicate_strategy}, 
                                                   request.remote_addr, request.user_agent.string)
                                                   
                    flash(f'✅ Successfully imported {success_count} members!', 'success')
                    return redirect(url_for('dashboard'))
                if error_count > 0:
                    flash(f'⚠️ {error_count} errors occurred.', 'error')
                    for error in errors[:5]:
                        flash(error, 'error')
                
                return redirect(url_for('bulk_import'))
                
            except Exception as e:
                flash(f'❌ Import failed: {str(e)}', 'error')
                return redirect(url_for('bulk_import'))
    
    # Clear any stale session data
    session.pop('import_file_path', None)
    session.pop('import_validation', None)
    
    return render_template('bulk_import.html')

@app.route('/import_preview')
def import_preview():
    """Show preview of import data with validation results"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    validation_data = session.get('import_validation')
    if not validation_data:
        flash('No import data found. Please upload a file first.', 'error')
        return redirect(url_for('bulk_import'))
    
    # Apply duplicate handling preview
    from import_validator import ImportValidator
    validator = ImportValidator(gym)
    
    # Get summary
    summary = validation_data['summary']
    duplicate_count = sum(1 for row in summary['rows'] if row.get('existing_member'))
    
    return render_template('import_preview.html',
                         validation=summary,
                         filename=validation_data['filename'],
                         duplicate_count=duplicate_count)

@app.route('/download_template')
def download_template():
    """Download sample Excel template for bulk import - USING OPENPYXL (NO PANDAS)"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    
    # Headers
    headers = [
        'Name', 'Phone', 'Email', 'Membership Type', 'Joined Date',
        'Status', 'Paid Month', 'Amount', 'Paid Months', 'Amounts'
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    center_align = Alignment(horizontal='center')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Sample Data (with Payment Example)
    current_month = datetime.now().strftime('%Y-%m')
    sample_data = [
        ['John Doe', '03001234567', 'john@example.com', 'Gym', '2025-01-01', 'Paid', current_month, 2500, '', ''],
        ['Jane Smith', '03117654321', 'jane@example.com', 'Gym + Cardio', '2025-01-05', 'Unpaid', '', '', '', ''],
        ['Ahmed Ali', '03009876543', 'ahmed@example.com', 'Gym', '2025-01-10', 'Paid', '', '', '2024-11,2024-12,2025-01', '2500,2500,3000']
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = cell_value
    
    # Adjust column widths
    column_widths = [20, 15, 25, 20, 15, 12, 12, 10, 24, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Save to buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, download_name='member_import_template.xlsx', as_attachment=True)

@app.route('/member/<member_id>/wallet_pass')
def generate_wallet_pass(member_id):
    """Generate Google Wallet pass for member"""
    gym = get_gym()
    if not gym: return redirect(url_for('auth'))
    
    member = gym.get_member(member_id)
    if not member:
        flash('Member not found!', 'error')
        return redirect(url_for('dashboard'))
    
    # Initialize wallet pass generator
    wallet = GymWalletPass()
    
    if not wallet.is_configured():
        flash('⚠️ Google Wallet not configured. Contact admin to set up.', 'error')
        return redirect(url_for('member_details', member_id=member_id))
    
    # Get gym details
    gym_details = gym.get_gym_details()
    
    # Create or update the loyalty class (one per gym)
    class_id = wallet.create_class(
        gym_name=gym_details['name'],
        gym_logo_url=None  # Can add logo URL later
    )
    
    if not class_id:
        flash('❌ Failed to create wallet class. Check credentials.', 'error')
        return redirect(url_for('member_details', member_id=member_id))
    
    # Generate the "Add to Google Wallet" URL
    save_url = wallet.create_jwt_save_url(
        member_id=member_id,
        member_name=member['name'],
        member_phone=member['phone'],
        gym_name=gym_details['name']
    )
    
    if save_url:
        # Redirect to Google Wallet
        return redirect(save_url)
    else:
        flash('❌ Failed to generate wallet pass. Check configuration.', 'error')
        return redirect(url_for('member_details', member_id=member_id))




@app.route('/fix_db')
def fix_database_schema():
    """Bulletproof database migration - CANNOT FAIL"""
    from sqlalchemy import text
    import traceback
    
    results = []
    
    try:
        from models import get_session
        session = get_session()
        
        # Wrap EVERYTHING in try-catch
        operations = [
            ("CREATE TABLE body_measurements", """
                CREATE TABLE IF NOT EXISTS body_measurements (
                    id SERIAL PRIMARY KEY,
                    member_id INTEGER REFERENCES members(id),
                    weight FLOAT,
                    body_fat FLOAT,
                    chest FLOAT,
                    waist FLOAT,
                    arms FLOAT,
                    notes TEXT,
                    recorded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """),
            ("CREATE TABLE member_notes", """
                CREATE TABLE IF NOT EXISTS member_notes (
                    id SERIAL PRIMARY KEY,
                    member_id INTEGER REFERENCES members(id),
                    note TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """),
            ("ADD members.birthday", "ALTER TABLE members ADD COLUMN IF NOT EXISTS birthday DATE"),
            ("ADD members.last_check_in", "ALTER TABLE members ADD COLUMN IF NOT EXISTS last_check_in TIMESTAMP"),
            ("ADD attendance.created_at", "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
            ("ADD attendance.emotion", "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS emotion VARCHAR(50)"),
            ("ADD attendance.confidence", "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS confidence FLOAT"),
            ("ADD users.market", "ALTER TABLE users ADD COLUMN IF NOT EXISTS market VARCHAR(50) DEFAULT 'US'"),
            ("ADD users.subscription_expiry", "ALTER TABLE users ADD COLUMN IF NOT EXISTS subscription_expiry TIMESTAMP"),
        ]
        
        for name, sql in operations:
            try:
                session.execute(text(sql))
                session.commit()
                results.append(f"✅ {name}")
            except Exception as e:
                session.rollback()
                error_msg = str(e)
                if "already exists" in error_msg or "duplicate" in error_msg.lower():
                    results.append(f"ℹ️ {name} (already exists)")
                else:
                    results.append(f"⚠️ {name}: {error_msg[:100]}")
        
        session.close()
        
        # Build success HTML
        html = """
        <html>
        <head>
            <title>Database Fixed</title>
            <style>
                body {
                    font-family: 'Inter', Arial, sans-serif;
                    background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
                    color: white;
                    padding: 2rem;
                    margin: 0;
                }
                .container {
                    max-width: 800px;
                    margin: 0 auto;
                    background: rgba(30, 41, 59, 0.8);
                    padding: 2.5rem;
                    border-radius: 16px;
                    box-shadow: 0 20px 60px rgba(0, 0, 0, 0.5);
                }
                h1 {
                    color: #8b5cf6;
                    margin-bottom: 2rem;
                    font-size: 2rem;
                }
                .result {
                    padding: 0.75rem 1rem;
                    margin: 0.5rem 0;
                    border-radius: 8px;
                    background: rgba(139, 92, 246, 0.1);
                    border-left: 4px solid #8b5cf6;
                }
                .success { border-left-color: #10b981; background: rgba(16, 185, 129, 0.1); }
                .info { border-left-color: #06b6d4; background: rgba(6, 182, 212, 0.1); }
                .warning { border-left-color: #f59e0b; background: rgba(245, 158, 11, 0.1); }
                .btn {
                    display: inline-block;
                    padding: 1rem 2rem;
                    background: linear-gradient(135deg, #8b5cf6, #06b6d4);
                    color: white;
                    text-decoration: none;
                    border-radius: 12px;
                    font-weight: 600;
                    margin-top: 2rem;
                    transition: transform 0.2s;
                }
                .btn:hover { transform: translateY(-2px); }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🔧 Database Migration Complete!</h1>
        """
        
        for result in results:
            css_class = "success" if "✅" in result else ("info" if "ℹ️" in result else "warning")
            html += f'<div class="result {css_class}">{result}</div>'
        
        html += """
                <a href="/" class="btn">→ Go to Dashboard</a>
            </div>
        </body>
        </html>
        """
        
        return html
        
    except Exception as e:
        # Even if EVERYTHING fails, show helpful error
        return f"""
        <html>
        <head><title>Migration Error</title></head>
        <body style='font-family: Arial; padding: 2rem; background: #1e293b; color: white;'>
            <h1 style='color: #ef4444;'>❌ Database Migration Error</h1>
            <div style='background: rgba(239, 68, 68, 0.1); padding: 1.5rem; border-radius: 12px; border-left: 4px solid #ef4444;'>
                <strong>Error:</strong><br>
                <pre style='color: #fca5a5; margin-top: 1rem;'>{str(e)}</pre>
                <hr style='border-color: rgba(239, 68, 68, 0.3); margin: 1rem 0;'>
                <strong>Traceback:</strong><br>
                <pre style='color: #fca5a5; font-size: 0.85rem;'>{traceback.format_exc()}</pre>
            </div>
            <a href='/' style='display: inline-block; margin-top: 2rem; padding: 1rem 2rem; background: #06b6d4; color: white; text-decoration: none; border-radius: 8px;'>Try Dashboard Anyway</a>
        </body>
        </html>
        """

@app.route('/api/chatbot', methods=['POST'])
def chatbot_api():
    """AI Chatbot API - Smart responses for gym queries"""
    try:
        data = request.get_json()
        message = data.get('message', '').lower().strip()
        
        if not message:
            return jsonify({'error': 'No message provided'}), 400
        
        # Get gym details for personalized responses
        username = session.get('username')
        gym_name = "Gym Manager"
        
        if username:
            try:
                gym_manager = GymManager(username)
                gym_details = gym_manager.get_gym_details()
                gym_name = gym_details.get('name', 'Gym Manager')
            except:
                pass
        
        # Smart response logic
        response = generate_smart_response(message, gym_name, username)
        
        return jsonify({
            'response': response['text'],
            'quick_replies': response.get('quick_replies', [])
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def generate_smart_response(message, gym_name, username=None):
    """Generate intelligent responses based on message content with a VIP brand voice"""
    
    # Greeting / Start
    if any(word in message for word in ['hi', 'hello', 'hey', 'start']):
        return {
            'text': f"💎 **Welcome to {gym_name} VIP Concierge!**\n\n"
                   f"How can I assist you today? I can help you with **Subscriptions**, **Gym Hours**, **Classes**, or **Account** details.\n\n"
                   "What's on your mind? 🔥",
            'quick_replies': ['Subscription Plans', 'Gym Hours', 'Our Facilities', 'Contact Us']
        }
    
    # Subscription / Plans (Dedicated Block)
    elif any(word in message for word in ['subscription', 'sub', 'membership', 'tier', 'plan', 'package', 'pricing', 'price']):
        return {
            'text': " ✨ **Exclusive Membership Tiers** ✨\n\n"
                   "Choose a plan that fits your fitness journey. You can view all details and **upgrade instantly** here:\n"
                   "🔗 **[View & Upgrade Subscription Plans](/subscription_plans)**\n\n"
                   "🏋️ **Starter VIP** — Rs 2,500/mo\n"
                   "💪 **Professional VIP** — Rs 4,500/mo\n"
                   "👑 **Ultimate Elite** — Rs 7,500/mo\n\n"
                   "💡 *Pay yearly to save 20%! Click the link above to get started.*",
            'quick_replies': ['Payment Options', 'Gym Hours', 'Book a Tour']
        }
    
    # Hours / Timing
    elif any(word in message for word in ['hour', 'time', 'open', 'close', 'timing', 'schedule']):
        return {
            'text': f"🕒 **{gym_name} Operational Hours:**\n\n"
                   "━━━━━━━━━━━━━━\n"
                   "📅 **Mon - Fri:** 6:00 AM - 11:00 PM\n"
                   "📅 **Sat - Sun:** 7:00 AM - 9:00 PM\n"
                   "━━━━━━━━━━━━━━\n\n"
                   "We are open 365 days a year to keep you consistent! 💪",
            'quick_replies': ['Subscription Plans', 'Our Facilities']
        }
    
    # Payment / Bills
    elif any(word in message for word in ['payment', 'pay', 'due', 'bill', 'fee', 'charge', 'invoice', 'gateway']):
        return {
            'text': "💳 **VIP Payment Portal Info**\n\n"
                   "We offer multiple convenient ways to pay. Manage your billing details here:\n"
                   "🔗 **[Go to Payment Settings](/settings)**\n\n"
                   "✅ **Direct Transfer:** JazzCash / EasyPaisa\n"
                   "✅ **Digital Card:** Visa, Mastercard, Amex\n"
                   "✅ **At Desk:** Cash or Card swipe\n\n"
                   "📍 Payments are due by the **5th** of each month.\n"
                   "💡 *Click 'Go to Payment Settings' to save your card for auto-pay.*",
            'quick_replies': ['Subscription Plans', 'Check My Balance', 'Contact Support']
        }
    
    # Facilities / Equipment
    elif any(word in message for word in ['facility', 'equipment', 'amenity', 'locker', 'spa', 'shower']):
        return {
            'text': "🏢 **Premium Amenities & Equipment**\n\n"
                   "Our facility is equipped with top-of-the-line gear:\n"
                   "🚀 **Strength Room:** Hammer Strength & Rogue rigs\n"
                   "🏃 **Cardio Suite:** Peloton bikes & Technogym treadmills\n"
                   "🧘 **Studio:** Yoga, Pilates, & HIIT zones\n"
                   "🚿 **Lounge:** Luxury showers & spa recovery\n\n"
                   "Experience the best in the city! 💫",
            'quick_replies': ['View Schedule', 'Subscription Plans']
        }
    
    # Contact / Support
    elif any(word in message for word in ['contact', 'phone', 'call', 'support', 'help', 'whatsapp', 'address']):
        return {
            'text': "📞 **Get in Touch with VIP Support**\n\n"
                   "We are here to assist you 24/7:\n\n"
                   "📱 **WhatsApp:** +92 300 1234567\n"
                   "📞 **Phone:** +92 300 1234567\n"
                   "✉️ **Email:** support@fitnessmanagement.site\n\n"
                   "📍 **Visit Us:** Main Boulevard, Fitness Plaza, Lahore.\n\n"
                   "Response time: Within 15 minutes! ⚡",
            'quick_replies': ['Subscription Plans', 'Gym Hours']
        }
    
    # Default Response
    else:
        return {
            'text': "🤖 **I'm your Fitness HQ AI.**\n\n"
                   "I'm here to help you dominate your goals. Ask me about:\n\n"
                   "⭐ **Membership Options**\n"
                   "⏰ **Timing & Schedules**\n"
                   "💳 **Billing & Payments**\n"
                   "🏋️ **Facilities & Trainers**\n\n"
                   "How can I serve you today?",
            'quick_replies': ['Subscription Plans', 'Gym Hours', 'Contact Us']
        }

@app.errorhandler(500)
def handle_internal_server_error(error):
    """Graceful fallback for unexpected server errors."""
    try:
        print("\n❌ INTERNAL SERVER ERROR")
        print(f"Path: {request.path}")
        print(traceback.format_exc())
    except Exception:
        pass

    # API-style endpoints should return JSON
    if request.path.startswith('/webhooks') or request.path.startswith('/stripe/webhook') or request.path.startswith('/chatbot_api'):
        return jsonify({'success': False, 'error': 'Internal server error'}), 500

    flash('Something went wrong on this page. Please try again.', 'error')

    # For authenticated users, keep UX smooth by sending them to dashboard
    if session.get('logged_in'):
        return redirect(url_for('dashboard'))

    return redirect(url_for('auth'))

if __name__ == '__main__':
    # Modular routes are now initialized at the global scope for Gunicorn compatibility
    
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
