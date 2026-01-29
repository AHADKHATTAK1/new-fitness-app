"""
Advanced Data Export Manager
Professional Excel/CSV exports with comprehensive member, payment, and analytics data
"""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta


class ExportManager:
    """Advanced export functionality for gym data"""
    
    def __init__(self, gym_manager):
        self.gym = gym_manager
    
    def export_members_complete(self) -> BytesIO:
        """
        Export complete member database with all details
        Includes: Personal info, membership details, payment history, attendance stats
        """
        members = self.gym.get_all_members()
        current_month = datetime.now().strftime('%Y-%m')
        
        data = []
        for member in members:
            # Get payment history
            fees = self.gym.get_member_fees(member['id'])
            total_paid = sum([float(fee.get('amount', 0)) for fee in fees])
            last_payment = fees[0].get('paid_date', 'Never') if fees else 'Never'
            
            # Payment status
            is_paid = self.gym.is_fee_paid(member['id'], current_month)
            
            data.append({
                'Member ID': member['id'],
                'Name': member['name'],
                'Phone': member['phone'],
                'Email': member.get('email', 'N/A'),
                'Join Date': member.get('joined_date', 'N/A'),
                'Membership Type': member.get('membership_type', 'Gym'),
                'Status': 'Active' if member.get('active', True) else 'Inactive',
                'Payment Status': 'PAID' if is_paid else 'UNPAID',
                'Last Payment': last_payment,
                'Total Payments': f"{self.gym.gym.currency}{total_paid:.2f}",
                'Is Trial': 'Yes' if member.get('is_trial') else 'No',
                'Trial End': member.get('trial_end_date', 'N/A'),
                'Birthday': member.get('birthday', 'N/A'),
                'Gender': member.get('gender', 'N/A'),
                'Height (cm)': member.get('height', 'N/A'),
                'Weight (kg)': member.get('weight', 'N/A'),
            })
        
        return self._create_styled_excel(data, 'Complete Members Database')
    
    def export_revenue_report(self, start_date: str = None, end_date: str = None) -> BytesIO:
        """
        Export detailed revenue report
        Includes: Daily/Monthly breakdown, payment methods, member types
        """
        # Get all fees in date range
        if not start_date:
            start_date = (datetime.now() - timedelta(days=90)).strftime('%Y-%m-%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y-%m-%d')
        
        # Query fees
        from models import Fee, Member
        fees_query = self.gym.session.query(Fee, Member).join(Member).filter(
            Member.gym_id == self.gym.gym.id,
            Fee.paid_date >= start_date,
            Fee.paid_date <= end_date
        ).all()
        
        data = []
        for fee, member in fees_query:
            data.append({
                'Date': fee.paid_date.strftime('%Y-%m-%d'),
                'Month': fee.month,
                'Member ID': member.id,
                'Member Name': member.name,
                'Phone': member.phone,
                'Membership Type': member.membership_type or 'Standard',
                'Amount': float(fee.amount),
                'Currency': self.gym.gym.currency,
            })
        
        return self._create_styled_excel(data, f'Revenue Report {start_date} to {end_date}')
    
    def export_attendance_analysis(self) -> BytesIO:
        """
        Export attendance analysis
        Includes: Check-in frequency, last visit, average visits per week
        """
        members = self.gym.get_all_members()
        
        data = []
        for member in members:
            # Get attendance records (if implemented)
            # For now, use placeholder logic
            data.append({
                'Member ID': member['id'],
                'Name': member['name'],
                'Phone': member['phone'],
                'Membership Type': member.get('membership_type', 'Gym'),
                'Join Date': member.get('joined_date', 'N/A'),
                'Status': 'Active' if member.get('active', True) else 'Inactive',
                # Placeholder attendance data
                'Total Visits': 'N/A',
                'Last Visit': 'N/A',
                'Avg Visits/Week': 'N/A',
            })
        
        return self._create_styled_excel(data, 'Attendance Analysis')
    
    def export_unpaid_members(self) -> BytesIO:
        """Export list of members with pending payments"""
        current_month = datetime.now().strftime('%Y-%m')
        members = self.gym.get_all_members()
        
        data = []
        for member in members:
            is_paid = self.gym.is_fee_paid(member['id'], current_month)
            if not is_paid and member.get('active', True):
                fees = self.gym.get_member_fees(member['id'])
                last_payment = fees[0].get('paid_date', 'Never') if fees else 'Never'
                
                data.append({
                    'Member ID': member['id'],
                    'Name': member['name'],
                    'Phone': member['phone'],
                    'Email': member.get('email', 'N/A'),
                    'Last Payment': last_payment,
                    'Months Unpaid': self._calculate_months_unpaid(fees, current_month),
                    'Amount Due': f"{self.gym.gym.currency}{self._calculate_due_amount(member)}",
                })
        
        return self._create_styled_excel(data, 'Unpaid Members Report')
    
    def _calculate_months_unpaid(self, fees, current_month):
        """Calculate how many months member hasn't paid"""
        if not fees:
            return "All time"
        
        last_fee = fees[0]
        last_month = datetime.strptime(last_fee['month'], '%Y-%m')
        current = datetime.strptime(current_month, '%Y-%m')
        
        months_diff = (current.year - last_month.year) * 12 + current.month - last_month.month
        return max(1, months_diff)
    
    def _calculate_due_amount(self, member):
        """Calculate total amount due (simplified - use actual fee structure)"""
        return 5000  # Placeholder
    
    def _create_styled_excel(self, data: list, sheet_name: str) -> BytesIO:
        """
        Create professionally styled Excel file
        """
        if not data:
            # Create empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            ws['A1'] = 'No data available'
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel sheet name limit
        
        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Header styling
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Border
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
